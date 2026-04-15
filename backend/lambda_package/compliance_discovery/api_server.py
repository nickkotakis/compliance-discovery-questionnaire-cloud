"""Flask REST API server for compliance discovery questionnaire."""

from flask import Flask, jsonify, request
from flask_cors import CORS
from typing import Dict, Any, List
from datetime import datetime
import json
import os

from compliance_discovery.nist_parser import NIST80053Parser
from compliance_discovery.csf_parser import NISTCSFParser
from compliance_discovery.cmmc_parser import CMMCParser
from compliance_discovery.question_generator import DiscoveryQuestionGenerator
from compliance_discovery.mcp_integration import MCPClient, create_aws_hints
from compliance_discovery.models.control import Control
from compliance_discovery.models.question import DiscoveryQuestion
from compliance_discovery.database import db, SessionModel
from compliance_discovery.control_descriptions import get_control_description
from compliance_discovery.aws_control_mapping import get_aws_responsibility, get_aws_services
from compliance_discovery.framework_mapper import get_framework_relevance
from compliance_discovery.csf_control_mapping import get_csf_responsibility, get_csf_aws_services
from compliance_discovery.csf_framework_mapper import get_csf_framework_relevance
from compliance_discovery.csf_organizational_controls import get_organizational_requirements, get_category_metadata
from compliance_discovery.cmmc_organizational_controls import get_cmmc_organizational_requirements, get_cmmc_category_metadata
from compliance_discovery.nist_800_53_organizational_controls import get_nist_organizational_requirements, get_nist_category_metadata
from compliance_discovery.cmmc_responsibility import get_cmmc_responsibility
from compliance_discovery.preventive_controls import get_preventive_controls_for_control, get_control_type_label
from compliance_discovery.ai_assistant import invoke_assistant


app = Flask(__name__)
CORS(app, resources={
    r"/api/*": {
        "origins": "*",
        "methods": ["GET", "POST", "PUT", "DELETE", "OPTIONS"],
        "allow_headers": ["Content-Type"]
    }
})

# Base directory for data files
basedir = os.path.abspath(os.path.dirname(__file__))

# Database configuration - use /tmp for Lambda (only writable directory)
if os.environ.get('AWS_LAMBDA_FUNCTION_NAME'):
    db_path = '/tmp/sessions.db'
else:
    db_path = os.path.join(basedir, 'sessions.db')

app.config['SQLALCHEMY_DATABASE_URI'] = f'sqlite:///{db_path}'
app.config['SQLALCHEMY_TRACK_MODIFICATIONS'] = False

# Initialize database
db.init_app(app)

# Create tables on startup
with app.app_context():
    db.create_all()

# Global state - keyed by framework
controls_cache: Dict[str, List[Control]] = {}  # framework -> controls
questions_cache: Dict[str, Dict[str, List[DiscoveryQuestion]]] = {}  # framework -> {control_id -> questions}
mcp_data_cache: Dict[str, List[Dict[str, Any]]] = {}  # Cache for MCP data from JSON file
csf_aws_data_cache: Dict[str, List[Dict[str, Any]]] = {}  # Cache for CSF AWS mappings
cmmc_aws_data_cache: Dict[str, List[Dict[str, Any]]] = {}  # Cache for CMMC AWS mappings

# Initialize components
parser = NIST80053Parser()
csf_parser = NISTCSFParser()
cmmc_parser = CMMCParser()
generator = DiscoveryQuestionGenerator()
mcp_client = MCPClient()

# Supported frameworks
SUPPORTED_FRAMEWORKS = {
    'nist-800-53': 'NIST 800-53 Rev 5 Moderate Baseline',
    'nist-csf': 'NIST CSF 2.0',
    'cmmc': 'CMMC Level 2 (v2.0)'
}

# CSF Function display names
CSF_FUNCTION_NAMES = {
    'GV': 'Govern',
    'ID': 'Identify',
    'PR': 'Protect',
    'DE': 'Detect',
    'RS': 'Respond',
    'RC': 'Recover'
}

# CMMC Domain display names
CMMC_DOMAIN_NAMES = {
    'AC': 'Access Control',
    'AT': 'Awareness and Training',
    'AU': 'Audit and Accountability',
    'CM': 'Configuration Management',
    'IA': 'Identification and Authentication',
    'IR': 'Incident Response',
    'MA': 'Maintenance',
    'MP': 'Media Protection',
    'PE': 'Physical Protection',
    'PS': 'Personnel Security',
    'RA': 'Risk Assessment',
    'CA': 'Security Assessment',
    'SC': 'System and Communications Protection',
    'SI': 'System and Information Integrity'
}


def load_mcp_data():
    """Load AWS controls data from MCP JSON file."""
    global mcp_data_cache
    
    mcp_file = os.path.join(basedir, 'aws_controls_mcp_data.json')
    if os.path.exists(mcp_file):
        try:
            with open(mcp_file, 'r') as f:
                data = json.load(f)
                mcp_data_cache = data.get('controls', {})
                print(f"Loaded MCP data for {len(mcp_data_cache)} controls from {mcp_file}")
        except Exception as e:
            print(f"Warning: Could not load MCP data file: {str(e)}")
            mcp_data_cache = {}
    else:
        print(f"Note: MCP data file not found at {mcp_file}")
        mcp_data_cache = {}


def load_csf_aws_data():
    """Load AWS controls data mapped to CSF subcategories."""
    global csf_aws_data_cache
    
    csf_aws_file = os.path.join(basedir, 'csf_aws_mappings.json')
    if os.path.exists(csf_aws_file):
        try:
            with open(csf_aws_file, 'r') as f:
                data = json.load(f)
                csf_aws_data_cache = data.get('controls', {})
                print(f"Loaded CSF AWS data for {len(csf_aws_data_cache)} subcategories from {csf_aws_file}")
        except Exception as e:
            print(f"Warning: Could not load CSF AWS data file: {str(e)}")
            csf_aws_data_cache = {}
    else:
        print(f"Note: CSF AWS data file not found at {csf_aws_file}")
        csf_aws_data_cache = {}


def load_cmmc_aws_data():
    """Load AWS controls data mapped to CMMC practices."""
    global cmmc_aws_data_cache
    
    cmmc_aws_file = os.path.join(basedir, 'cmmc_aws_mappings.json')
    if os.path.exists(cmmc_aws_file):
        try:
            with open(cmmc_aws_file, 'r') as f:
                data = json.load(f)
                cmmc_aws_data_cache = data.get('controls', {})
                print(f"Loaded CMMC AWS data for {len(cmmc_aws_data_cache)} practices from {cmmc_aws_file}")
        except Exception as e:
            print(f"Warning: Could not load CMMC AWS data file: {str(e)}")
            cmmc_aws_data_cache = {}
    else:
        print(f"Note: CMMC AWS data file not found at {cmmc_aws_file}")
        cmmc_aws_data_cache = {}


def initialize_data(framework: str = 'nist-800-53'):
    """Initialize controls and questions data for a given framework.
    
    Args:
        framework: Framework identifier ('nist-800-53' or 'nist-csf')
    """
    global controls_cache, questions_cache
    
    if framework not in controls_cache or not controls_cache[framework]:
        # Load MCP data first (shared across frameworks)
        if not mcp_data_cache:
            load_mcp_data()
        
        if framework == 'nist-800-53':
            print("Loading NIST 800-53 Moderate Baseline controls...")
            controls_cache[framework] = parser.get_moderate_baseline_controls()
            print(f"Loaded {len(controls_cache[framework])} controls")
            
            # Pass AWS controls data to generator
            generator.set_aws_controls_data(mcp_data_cache)
            
            print("Generating discovery questions for NIST 800-53...")
            questions_cache[framework] = {}
            for control in controls_cache[framework]:
                aws_controls = mcp_data_cache.get(control.id.lower(), [])
                questions = generator.generate_questions(control, aws_controls)
                questions_cache[framework][control.id] = questions
            print(f"Generated questions for {len(questions_cache[framework])} controls")
            
        elif framework == 'nist-csf':
            print("Loading NIST CSF 2.0 subcategories...")
            controls_cache[framework] = csf_parser.get_all_subcategories()
            print(f"Loaded {len(controls_cache[framework])} subcategories")
            
            # Load CSF-specific AWS mappings
            if not csf_aws_data_cache:
                load_csf_aws_data()
            
            print("Generating discovery questions for NIST CSF 2.0...")
            questions_cache[framework] = {}
            for control in controls_cache[framework]:
                aws_controls = csf_aws_data_cache.get(control.id.lower(), [])
                questions = generator.generate_csf_questions(control, aws_controls)
                questions_cache[framework][control.id] = questions
            print(f"Generated questions for {len(questions_cache[framework])} subcategories")

        elif framework == 'cmmc':
            print("Loading CMMC Level 2 practices...")
            controls_cache[framework] = cmmc_parser.get_all_practices()
            print(f"Loaded {len(controls_cache[framework])} practices")
            
            # Load CMMC-specific AWS mappings
            if not cmmc_aws_data_cache:
                load_cmmc_aws_data()
            
            print("Generating discovery questions for CMMC Level 2...")
            questions_cache[framework] = {}
            for control in controls_cache[framework]:
                aws_controls = cmmc_aws_data_cache.get(control.id.lower(), [])
                questions = generator.generate_cmmc_questions(control, aws_controls)
                questions_cache[framework][control.id] = questions
            print(f"Generated questions for {len(questions_cache[framework])} practices")


@app.after_request
def after_request(response):
    """Add CORS headers to all responses."""
    response.headers.add('Access-Control-Allow-Origin', '*')
    response.headers.add('Access-Control-Allow-Headers', 'Content-Type')
    response.headers.add('Access-Control-Allow-Methods', 'GET,POST,PUT,DELETE,OPTIONS')
    return response


@app.route('/api/frameworks', methods=['OPTIONS'])
@app.route('/api/controls', methods=['OPTIONS'])
@app.route('/api/controls/<control_id>', methods=['OPTIONS'])
@app.route('/api/questions', methods=['OPTIONS'])
@app.route('/api/session', methods=['OPTIONS'])
@app.route('/api/session/<session_id>', methods=['OPTIONS'])
@app.route('/api/session/<session_id>/response', methods=['OPTIONS'])
@app.route('/api/sessions', methods=['OPTIONS'])
@app.route('/api/export', methods=['OPTIONS'])
def handle_options(control_id=None, session_id=None):
    """Handle OPTIONS preflight requests."""
    return '', 204


@app.route('/api/health', methods=['GET'])
def health_check():
    """Health check endpoint."""
    return jsonify({
        'status': 'healthy',
        'timestamp': datetime.utcnow().isoformat()
    })


@app.route('/api/frameworks', methods=['GET'])
def get_frameworks():
    """Get available compliance frameworks."""
    frameworks = []
    for key, label in SUPPORTED_FRAMEWORKS.items():
        info = {'id': key, 'label': label}
        if key == 'nist-csf':
            info['functions'] = {
                code: name for code, name in CSF_FUNCTION_NAMES.items()
            }
        elif key == 'cmmc':
            info['domains'] = {
                code: name for code, name in CMMC_DOMAIN_NAMES.items()
            }
        frameworks.append(info)
    return jsonify({'frameworks': frameworks})


@app.route('/api/controls', methods=['GET'])
def get_controls():
    """Get all controls for a framework.
    
    Query parameters:
        framework: Framework ID ('nist-800-53' or 'nist-csf', default: 'nist-800-53')
        family: Filter by control family / CSF function (optional)
    """
    framework = request.args.get('framework', 'nist-800-53')
    if framework not in SUPPORTED_FRAMEWORKS:
        return jsonify({'error': f'Unsupported framework: {framework}'}), 400
    
    initialize_data(framework)
    
    family = request.args.get('family')
    
    cached = controls_cache.get(framework, [])
    filtered_controls = cached
    if family:
        filtered_controls = [c for c in cached if c.family.upper() == family.upper()]
    
    if framework == 'nist-csf':
        # Sort CSF subcategories by function then category then number
        def csf_sort_key(control):
            parts = control.id.split('-')
            prefix = parts[0] if parts else control.id  # e.g., "GV.OC"
            num = int(parts[1]) if len(parts) > 1 else 0
            func_order = {'GV': 0, 'ID': 1, 'PR': 2, 'DE': 3, 'RS': 4, 'RC': 5}
            func_code = control.family.upper()
            return (func_order.get(func_code, 99), prefix, num)
        
        sorted_controls = sorted(filtered_controls, key=csf_sort_key)
        
        return jsonify({
            'controls': [
                {
                    'id': c.id,
                    'title': c.title,
                    'description': c.description,
                    'family': c.family,
                    'function_name': CSF_FUNCTION_NAMES.get(c.family.upper(), c.family),
                    'category': c.id.split('-')[0] if '-' in c.id else c.id,
                    'category_name': csf_parser.get_category_name(c.id.split('-')[0] if '-' in c.id else c.id),
                    'in_moderate_baseline': True,
                    'aws_responsibility': get_csf_responsibility(c.id)
                }
                for c in sorted_controls
            ],
            'total': len(sorted_controls),
            'framework': framework,
            'framework_label': SUPPORTED_FRAMEWORKS[framework]
        })
    elif framework == 'cmmc':
        # Sort CMMC practices by domain then practice number
        CMMC_DOMAIN_ORDER = {
            'AC': 0, 'AT': 1, 'AU': 2, 'CM': 3, 'IA': 4, 'IR': 5, 'MA': 6,
            'MP': 7, 'PE': 8, 'PS': 9, 'RA': 10, 'CA': 11, 'SC': 12, 'SI': 13
        }

        def cmmc_sort_key(control):
            domain = control.family.upper()
            # Practice IDs like "AC.L2-3.1.1" — extract the trailing number for sorting
            parts = control.id.split('-')
            num_part = parts[-1] if len(parts) > 1 else '0'
            # Convert dotted number to tuple for proper sorting (e.g., "3.1.1" -> (3,1,1))
            try:
                num_tuple = tuple(int(x) for x in num_part.split('.'))
            except ValueError:
                num_tuple = (0,)
            return (CMMC_DOMAIN_ORDER.get(domain, 99), num_tuple)

        sorted_controls = sorted(filtered_controls, key=cmmc_sort_key)

        return jsonify({
            'controls': [
                {
                    'id': c.id,
                    'title': c.title,
                    'description': c.description,
                    'family': c.family,
                    'domain_name': CMMC_DOMAIN_NAMES.get(c.family.upper(), c.family),
                    'in_moderate_baseline': True,
                    'aws_responsibility': get_cmmc_responsibility(c.id)
                }
                for c in sorted_controls
            ],
            'total': len(sorted_controls),
            'framework': framework,
            'framework_label': SUPPORTED_FRAMEWORKS[framework]
        })
    else:
        # Original NIST 800-53 sorting
        def sort_key(control):
            parts = control.id.upper().split('-')
            if len(parts) == 2:
                family_code = parts[0]
                try:
                    if '(' in parts[1]:
                        base_num = int(parts[1].split('(')[0])
                        enh_num = int(parts[1].split('(')[1].rstrip(')'))
                        return (family_code, base_num, enh_num)
                    else:
                        return (family_code, int(parts[1]), 0)
                except ValueError:
                    return (family_code, 0, 0)
            return (control.id.upper(), 0, 0)
        
        sorted_controls = sorted(filtered_controls, key=sort_key)
        
        return jsonify({
            'controls': [
                {
                    'id': c.id,
                    'title': c.title,
                    'description': get_control_description(c.id) or c.description,
                    'family': c.family,
                    'in_moderate_baseline': c.in_moderate_baseline,
                    'parameter_count': len(c.parameters),
                    'enhancement_count': len(c.enhancements),
                    'aws_responsibility': get_aws_responsibility(c.id)
                }
                for c in sorted_controls
            ],
            'total': len(sorted_controls),
            'framework': framework,
            'framework_label': SUPPORTED_FRAMEWORKS[framework]
        })


@app.route('/api/controls/<control_id>', methods=['GET'])
def get_control(control_id: str):
    """Get specific control with questions and AWS mappings.
    
    Args:
        control_id: Control ID (e.g., "AC-1" or "GV.OC-01")
    
    Query parameters:
        framework: Framework ID ('nist-800-53' or 'nist-csf', default: 'nist-800-53')
    """
    framework = request.args.get('framework', 'nist-800-53')
    if framework not in SUPPORTED_FRAMEWORKS:
        return jsonify({'error': f'Unsupported framework: {framework}'}), 400
    
    initialize_data(framework)
    
    # Normalize control ID for comparison
    if framework == 'nist-csf':
        # CSF IDs are like GV.OC-01 — normalize to uppercase
        normalized_id = control_id.upper()
    else:
        normalized_id = control_id.upper()
    
    # Find control (case-insensitive)
    cached = controls_cache.get(framework, [])
    control = next((c for c in cached if c.id.upper() == normalized_id), None)
    if not control:
        print(f"Control not found: {control_id} (normalized: {normalized_id}) in framework {framework}")
        return jsonify({'error': f'Control not found: {control_id}'}), 404
    
    # Get questions
    fw_questions = questions_cache.get(framework, {})
    questions = fw_questions.get(control.id, [])
    
    # Try to get AWS mappings using normalized ID
    aws_hints = []
    aws_controls_data = []  # Store full control data for detailed view
    
    # Choose the right data cache based on framework
    if framework == 'nist-csf':
        data_cache = csf_aws_data_cache
    elif framework == 'cmmc':
        data_cache = cmmc_aws_data_cache
    else:
        data_cache = mcp_data_cache
    
    # First try loading from data cache (JSON file)
    if normalized_id.lower() in data_cache:
        aws_controls_data = data_cache[normalized_id.lower()]
        print(f"Loaded {len(aws_controls_data)} AWS controls for {normalized_id} from MCP cache")
        # Generate hints from cached data
        if aws_controls_data:
            for ac in aws_controls_data:
                hint_parts = []
                if ac.get('services'):
                    services = ", ".join(ac['services'][:2])
                    hint_parts.append(f"Services: {services}")
                if ac.get('config_rules'):
                    rules = ", ".join(ac['config_rules'][:2])
                    hint_parts.append(f"Config: {rules}")
                if ac.get('security_hub_controls'):
                    hub = ", ".join(ac['security_hub_controls'][:2])
                    hint_parts.append(f"Security Hub: {hub}")
                if hint_parts:
                    aws_hints.append(f"• {ac['title']}\n  {' | '.join(hint_parts)}")
    
    # Fallback: Try MCP server (if available)
    if not aws_controls_data:
        try:
            if mcp_client.connected or mcp_client.connect():
                aws_controls = mcp_client.map_compliance_requirements(normalized_id)
                if aws_controls:
                    aws_hints = create_aws_hints(aws_controls)
                    # Store detailed control data
                    aws_controls_data = [
                        {
                            'control_id': ac.control_id,
                            'title': ac.title,
                            'description': ac.description,
                            'services': ac.services,
                            'config_rules': ac.managed_controls.config_rules,
                            'security_hub_controls': ac.managed_controls.security_hub_controls,
                            'control_tower_ids': ac.managed_controls.control_tower_ids,
                            'frameworks': ac.frameworks
                        }
                        for ac in aws_controls
                    ]
        except Exception as e:
            print(f"Warning: Could not fetch AWS mappings from MCP: {str(e)}")
    
    # Final fallback to manual mapping if nothing else worked
    if not aws_hints:
        if framework == 'nist-csf':
            fallback_services = get_csf_aws_services(control.id)
        elif framework == 'cmmc':
            fallback_services = []  # No manual fallback for CMMC yet
        else:
            fallback_services = get_aws_services(control.id)
        if fallback_services:
            aws_hints = [f"AWS Services: {', '.join(fallback_services)}"]
    
    # Determine responsibility using framework-specific mapping
    if framework == 'nist-csf':
        responsibility = get_csf_responsibility(control.id)
    elif framework == 'cmmc':
        responsibility = get_cmmc_responsibility(control.id)
    else:
        responsibility = get_aws_responsibility(control.id)
    
    # Build applicability message based on responsibility
    if responsibility == 'aws':
        aws_applicability = {
            'applicable': False,
            'responsibility': 'aws',
            'message': '🟠 AWS RESPONSIBILITY: AWS handles this control for their infrastructure. Your AWS workloads automatically inherit this protection. Download compliance reports from AWS Artifact to demonstrate this to auditors.',
            'artifact_links': [
                {
                    'name': 'AWS Artifact (Compliance Reports)',
                    'url': 'https://console.aws.amazon.com/artifact/',
                    'description': 'Download reports showing AWS implementation of this control'
                },
                {
                    'name': 'AWS Data Center Security',
                    'url': 'https://aws.amazon.com/compliance/data-center/',
                    'description': 'Learn about AWS physical security measures'
                }
            ],
            'controls': []
        }
    elif responsibility == 'shared' and aws_hints:
        aws_applicability = {
            'applicable': True,
            'responsibility': 'shared',
            'message': f'🟢 SHARED RESPONSIBILITY: AWS provides the services listed below. You must configure and use them properly in your AWS environment to meet this requirement.',
            'artifact_links': [
                {
                    'name': 'AWS Artifact (Compliance Reports)',
                    'url': 'https://console.aws.amazon.com/artifact/',
                    'description': 'Download SOC, PCI, ISO reports showing AWS platform compliance'
                },
                {
                    'name': 'AWS Compliance Programs',
                    'url': 'https://aws.amazon.com/compliance/programs/',
                    'description': 'View AWS compliance certifications'
                }
            ],
            'controls': aws_hints
        }
    else:
        aws_applicability = {
            'applicable': False,
            'responsibility': 'customer',
            'message': '🔵 CUSTOMER RESPONSIBILITY: You must implement this control in your AWS environment using custom configurations, third-party tools, or application-level controls.',
            'artifact_links': [],
            'controls': []
        }
    
    # Clean the description for display - use human-friendly version if available
    human_description = get_control_description(control.id)
    cleaned_description = human_description if human_description else control.description
    
    # Get framework relevance
    if framework == 'nist-csf':
        framework_relevance = get_csf_framework_relevance(control.id)
    elif framework == 'cmmc':
        # CMMC maps to NIST 800-171 / 800-53 — provide basic relevance info
        framework_relevance = {
            'control_id': control.id,
            'family': control.family,
            'relevant_frameworks': ['NIST SP 800-171 Rev 2', 'NIST SP 800-53 Rev 5'],
            'notes': f'CMMC Level 2 practice {control.id} maps to NIST SP 800-171 Rev 2 requirements.',
            'specific_mappings': {},
            'has_specific_mappings': False
        }
    else:
        framework_relevance = get_framework_relevance(control.id)
    
    return jsonify({
        'control': {
            'id': control.id,
            'title': control.title,
            'description': cleaned_description,
            'family': control.family,
            'in_moderate_baseline': control.in_moderate_baseline,
            'parameters': [
                {
                    'id': p.id,
                    'label': p.label,
                    'description': p.description,
                    'constraints': p.constraints
                }
                for p in control.parameters
            ],
            'enhancements': [
                {
                    'id': e.id,
                    'title': e.title,
                    'description': e.description,
                    'in_moderate_baseline': e.in_moderate_baseline
                }
                for e in control.enhancements
            ]
        },
        'questions': [
            {
                'id': q.id,
                'control_id': q.control_id,
                'question_text': q.question_text,
                'question_type': q.question_type.value,
                'family': q.family,
                'aws_service_guidance': q.aws_service_guidance
            }
            for q in questions
        ],
        'aws_hints': aws_hints,
        'aws_applicability': aws_applicability,
        'aws_controls': aws_controls_data,  # Add detailed AWS control data
        'preventive_controls': get_preventive_controls_for_control(control.id, framework),  # SCPs + OPA
        'framework_relevance': framework_relevance,  # Add framework relevance
        'organizational_requirements': get_organizational_requirements(control.id) if framework == 'nist-csf' else get_cmmc_organizational_requirements(control.id) if framework == 'cmmc' else get_nist_organizational_requirements(control.id),
        'organizational_category_metadata': get_category_metadata() if framework == 'nist-csf' else get_cmmc_category_metadata() if framework == 'cmmc' else get_nist_category_metadata(),
        'domain_name': CMMC_DOMAIN_NAMES.get(control.family.upper(), control.family) if framework == 'cmmc' else None,
        'framework': framework,
        'framework_label': SUPPORTED_FRAMEWORKS.get(framework, framework)
    })


@app.route('/api/questions', methods=['GET'])
def get_questions():
    """Get all questions.
    
    Query parameters:
        framework: Framework ID (default: 'nist-800-53')
        control_id: Filter by control ID (optional)
        family: Filter by family (optional)
        question_type: Filter by question type (optional)
    """
    framework = request.args.get('framework', 'nist-800-53')
    initialize_data(framework)
    
    control_id = request.args.get('control_id')
    family = request.args.get('family')
    question_type = request.args.get('question_type')
    
    fw_questions = questions_cache.get(framework, {})
    all_questions = []
    for questions in fw_questions.values():
        all_questions.extend(questions)
    
    # Apply filters
    if control_id:
        all_questions = [q for q in all_questions if q.control_id == control_id]
    if family:
        all_questions = [q for q in all_questions if q.family == family]
    if question_type:
        all_questions = [q for q in all_questions if q.question_type.value == question_type]
    
    return jsonify({
        'questions': [
            {
                'id': q.id,
                'control_id': q.control_id,
                'question_text': q.question_text,
                'question_type': q.question_type.value,
                'family': q.family,
                'aws_service_guidance': q.aws_service_guidance
            }
            for q in all_questions
        ],
        'total': len(all_questions)
    })


@app.route('/api/session', methods=['POST'])
def create_session():
    """Create new assessment session.
    
    Request body:
        {
            "customer_name": str,
            "analyst_name": str,
            "frameworks": List[str]
        }
    """
    data = request.get_json()
    
    session_id = f"session-{datetime.utcnow().timestamp()}"
    
    session_data = {
        'id': session_id,
        'customer_name': data.get('customer_name', ''),
        'analyst_name': data.get('analyst_name', ''),
        'frameworks': data.get('frameworks', ['NIST 800-53']),
        'status': 'active',
        'responses': {},
        'evidence': {}
    }
    
    # Save to database
    session_model = SessionModel.from_dict(session_data)
    db.session.add(session_model)
    db.session.commit()
    
    return jsonify(session_model.to_dict()), 201


@app.route('/api/session/<session_id>', methods=['GET'])
def get_session(session_id: str):
    """Get session details."""
    session = SessionModel.query.get(session_id)
    if not session:
        return jsonify({'error': 'Session not found'}), 404
    
    return jsonify(session.to_dict())


@app.route('/api/session/<session_id>/response', methods=['POST'])
def record_response(session_id: str):
    """Record a response to a question.
    
    Request body:
        {
            "question_id": str,
            "answer": str,
            "notes": str (optional)
        }
    """
    session = SessionModel.query.get(session_id)
    if not session:
        return jsonify({'error': 'Session not found'}), 404
    
    data = request.get_json()
    question_id = data.get('question_id')
    
    # Update responses
    responses = json.loads(session.responses) if session.responses else {}
    responses[question_id] = {
        'answer': data.get('answer', ''),
        'notes': data.get('notes', ''),
        'timestamp': datetime.utcnow().isoformat()
    }
    
    session.responses = json.dumps(responses)
    session.updated_at = datetime.utcnow()
    db.session.commit()
    
    return jsonify({'success': True})


@app.route('/api/export', methods=['GET'])
def export_template():
    """Export questionnaire in various formats.
    
    Query parameters:
        format: Export format (json, excel, pdf, yaml) - default: json
        framework: Framework ID (default: 'nist-800-53')
        include_unanswered: Include unanswered questions (default: true)
        include_aws_hints: Include AWS implementation hints (default: true)
        include_framework_mappings: Include framework mappings (default: true)
    """
    framework = request.args.get('framework', 'nist-800-53')
    initialize_data(framework)
    
    export_format = request.args.get('format', 'json')
    include_unanswered = request.args.get('include_unanswered', 'true').lower() == 'true'
    include_aws_hints = request.args.get('include_aws_hints', 'true').lower() == 'true'
    include_framework_mappings = request.args.get('include_framework_mappings', 'true').lower() == 'true'
    
    fw_controls = controls_cache.get(framework, [])
    fw_questions = questions_cache.get(framework, {})
    
    # Build template data
    template_data = {
        'metadata': {
            'template_version': '1.0.0',
            'baseline_version': SUPPORTED_FRAMEWORKS.get(framework, framework),
            'export_date': datetime.utcnow().isoformat(),
            'total_control_count': len(fw_controls),
            'frameworks_included': [SUPPORTED_FRAMEWORKS.get(framework, framework), 'AWS']
        },
        'controls': [
            {
                'id': c.id,
                'title': c.title,
                'description': get_control_description(c.id) or c.description,
                'family': c.family,
                'aws_responsibility': get_aws_responsibility(c.id) if framework == 'nist-800-53' else get_csf_responsibility(c.id)
            }
            for c in fw_controls
        ],
        'questions': {
            control_id: [
                {
                    'id': q.id,
                    'question_text': q.question_text,
                    'question_type': q.question_type.value,
                    'aws_service_guidance': q.aws_service_guidance if include_aws_hints else None,
                    'response': ''
                }
                for q in questions
            ]
            for control_id, questions in fw_questions.items()
        }
    }
    
    # Add framework mappings if requested
    if include_framework_mappings and framework == 'nist-800-53':
        template_data['framework_mappings'] = {
            control.id: get_framework_relevance(control.id)
            for control in fw_controls
        }
    
    # Add AWS hints if requested
    if include_aws_hints:
        template_data['aws_hints'] = {}
        data_cache = csf_aws_data_cache if framework == 'nist-csf' else mcp_data_cache
        for control in fw_controls:
            normalized_id = control.id.lower()
            if normalized_id in data_cache:
                aws_controls = data_cache[normalized_id]
                hints = []
                for ac in aws_controls:
                    hint_parts = []
                    if ac.get('services'):
                        services = ", ".join(ac['services'][:2])
                        hint_parts.append(f"Services: {services}")
                    if ac.get('config_rules'):
                        rules = ", ".join(ac['config_rules'][:2])
                        hint_parts.append(f"Config: {rules}")
                    if hint_parts:
                        hints.append(f"{ac['title']}: {' | '.join(hint_parts)}")
                template_data['aws_hints'][control.id] = hints
    
    if export_format == 'json':
        return jsonify(template_data)
    elif export_format == 'yaml':
        try:
            import yaml
            yaml_content = yaml.dump(template_data, default_flow_style=False, sort_keys=False)
            return yaml_content, 200, {'Content-Type': 'application/x-yaml', 'Content-Disposition': 'attachment; filename=compliance-questionnaire.yaml'}
        except ImportError:
            return jsonify({'error': 'YAML export requires PyYAML package'}), 500
    elif export_format == 'excel':
        try:
            from openpyxl import Workbook
            from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
            from io import BytesIO
            
            wb = Workbook()
            
            # Remove default sheet
            if 'Sheet' in wb.sheetnames:
                wb.remove(wb['Sheet'])
            
            # Define color scheme
            header_fill = PatternFill(start_color='366092', end_color='366092', fill_type='solid')
            header_font = Font(color='FFFFFF', bold=True, size=11)
            
            # Family colors (alternating for visual grouping)
            family_colors = {
                'ac': 'E3F2FD',  # Light Blue
                'at': 'F3E5F5',  # Light Purple
                'au': 'E8F5E9',  # Light Green
                'ca': 'FFF3E0',  # Light Orange
                'cm': 'FCE4EC',  # Light Pink
                'cp': 'E0F2F1',  # Light Teal
                'ia': 'FFF9C4',  # Light Yellow
                'ir': 'FFEBEE',  # Light Red
                'ma': 'E8EAF6',  # Light Indigo
                'mp': 'F1F8E9',  # Light Lime
                'pe': 'FBE9E7',  # Light Deep Orange
                'pl': 'E0F7FA',  # Light Cyan
                'pm': 'F9FBE7',  # Light Lime
                'ps': 'EDE7F6',  # Light Deep Purple
                'pt': 'E1F5FE',  # Light Light Blue
                'ra': 'F3E5F5',  # Light Purple
                'sa': 'E8F5E9',  # Light Green
                'sc': 'FFF3E0',  # Light Orange
                'si': 'FCE4EC',  # Light Pink
                'sr': 'E0F2F1',  # Light Teal
            }
            
            # Responsibility colors
            responsibility_colors = {
                'aws': 'FFE0B2',      # Orange
                'shared': 'C8E6C9',   # Green
                'customer': 'BBDEFB', # Blue
                'unknown': 'F5F5F5'   # Gray
            }
            
            # Border style
            thin_border = Border(
                left=Side(style='thin', color='CCCCCC'),
                right=Side(style='thin', color='CCCCCC'),
                top=Side(style='thin', color='CCCCCC'),
                bottom=Side(style='thin', color='CCCCCC')
            )
            
            # ===== CONTROLS SHEET =====
            ws_controls = wb.create_sheet("Controls", 0)
            
            # Headers
            headers = ['Control ID', 'Title', 'Description', 'Family', 'AWS Responsibility']
            for col_num, header in enumerate(headers, 1):
                cell = ws_controls.cell(row=1, column=col_num)
                cell.value = header
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                cell.border = thin_border
            
            # Sort controls by family, then by control number
            def sort_key(control):
                parts = control['id'].upper().split('-')
                if len(parts) == 2:
                    family_code = parts[0]
                    try:
                        if '(' in parts[1]:
                            base_num = int(parts[1].split('(')[0])
                            enh_num = int(parts[1].split('(')[1].rstrip(')'))
                            return (family_code, base_num, enh_num)
                        else:
                            return (family_code, int(parts[1]), 0)
                    except ValueError:
                        return (family_code, 0, 0)
                return (control['id'].upper(), 0, 0)
            
            sorted_controls = sorted(template_data['controls'], key=sort_key)
            
            # Data rows with family-based color coding
            current_family = None
            for idx, control in enumerate(sorted_controls, start=2):
                family = control['family'].lower()
                
                # Add family separator row when family changes
                if current_family != family:
                    current_family = family
                    # Add a bold family header row
                    family_row = ws_controls.cell(row=idx, column=1)
                    family_full_name = {
                        'ac': 'Access Control', 'at': 'Awareness and Training',
                        'au': 'Audit and Accountability', 'ca': 'Assessment, Authorization, and Monitoring',
                        'cm': 'Configuration Management', 'cp': 'Contingency Planning',
                        'ia': 'Identification and Authentication', 'ir': 'Incident Response',
                        'ma': 'Maintenance', 'mp': 'Media Protection',
                        'pe': 'Physical and Environmental Protection', 'pl': 'Planning',
                        'pm': 'Program Management', 'ps': 'Personnel Security',
                        'pt': 'PII Processing and Transparency', 'ra': 'Risk Assessment',
                        'sa': 'System and Services Acquisition', 'sc': 'System and Communications Protection',
                        'si': 'System and Information Integrity', 'sr': 'Supply Chain Risk Management'
                    }.get(family, family.upper())
                    
                    family_row.value = f"═══ {family.upper()} - {family_full_name} ═══"
                    family_row.font = Font(bold=True, size=12, color='366092')
                    family_row.alignment = Alignment(horizontal='left', vertical='center')
                    ws_controls.merge_cells(f'A{idx}:E{idx}')
                    for col in range(1, 6):
                        ws_controls.cell(row=idx, column=col).fill = PatternFill(
                            start_color='E8EAF6', end_color='E8EAF6', fill_type='solid'
                        )
                    idx += 1
                
                # Control data
                family_fill = PatternFill(
                    start_color=family_colors.get(family, 'FFFFFF'),
                    end_color=family_colors.get(family, 'FFFFFF'),
                    fill_type='solid'
                )
                
                # Control ID
                cell = ws_controls.cell(row=idx, column=1)
                cell.value = control['id']
                cell.font = Font(bold=True)
                cell.fill = family_fill
                cell.border = thin_border
                cell.alignment = Alignment(horizontal='center', vertical='center')
                
                # Title
                cell = ws_controls.cell(row=idx, column=2)
                cell.value = control['title']
                cell.fill = family_fill
                cell.border = thin_border
                cell.alignment = Alignment(wrap_text=True, vertical='top')
                
                # Description
                cell = ws_controls.cell(row=idx, column=3)
                cell.value = control['description']
                cell.fill = family_fill
                cell.border = thin_border
                cell.alignment = Alignment(wrap_text=True, vertical='top')
                
                # Family
                cell = ws_controls.cell(row=idx, column=4)
                cell.value = family.upper()
                cell.fill = family_fill
                cell.border = thin_border
                cell.alignment = Alignment(horizontal='center', vertical='center')
                
                # AWS Responsibility with color coding
                cell = ws_controls.cell(row=idx, column=5)
                responsibility = control.get('aws_responsibility', 'unknown')
                cell.value = responsibility.upper()
                cell.fill = PatternFill(
                    start_color=responsibility_colors.get(responsibility, 'F5F5F5'),
                    end_color=responsibility_colors.get(responsibility, 'F5F5F5'),
                    fill_type='solid'
                )
                cell.border = thin_border
                cell.alignment = Alignment(horizontal='center', vertical='center')
                cell.font = Font(bold=True)
            
            # Set column widths
            ws_controls.column_dimensions['A'].width = 15
            ws_controls.column_dimensions['B'].width = 40
            ws_controls.column_dimensions['C'].width = 60
            ws_controls.column_dimensions['D'].width = 12
            ws_controls.column_dimensions['E'].width = 18
            
            # Freeze header row
            ws_controls.freeze_panes = 'A2'
            
            # ===== QUESTIONS SHEET =====
            ws_questions = wb.create_sheet("Questions")
            
            # Headers
            question_headers = ['Control ID', 'Family', 'Question #', 'Question Text', 'Question Type', 'Response']
            for col_num, header in enumerate(question_headers, 1):
                cell = ws_questions.cell(row=1, column=col_num)
                cell.value = header
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                cell.border = thin_border
            
            # Question rows sorted by control (which are already sorted by family)
            row = 2
            current_family = None
            
            # Question type sort order (implementation first, then evidence, etc.)
            question_type_order = {
                'implementation': 1,
                'evidence': 2,
                'policy': 3,
                'procedure': 4,
                'technical': 5,
                'administrative': 6,
                'physical': 7
            }
            
            for control in sorted_controls:
                control_id = control['id']
                family = control['family'].lower()
                
                if control_id in template_data['questions']:
                    questions = template_data['questions'][control_id]
                    
                    # Sort questions by type
                    sorted_questions = sorted(
                        questions,
                        key=lambda q: question_type_order.get(q['question_type'].lower(), 99)
                    )
                    
                    # Add family separator when family changes
                    if current_family != family:
                        current_family = family
                        family_full_name = {
                            'ac': 'Access Control', 'at': 'Awareness and Training',
                            'au': 'Audit and Accountability', 'ca': 'Assessment, Authorization, and Monitoring',
                            'cm': 'Configuration Management', 'cp': 'Contingency Planning',
                            'ia': 'Identification and Authentication', 'ir': 'Incident Response',
                            'ma': 'Maintenance', 'mp': 'Media Protection',
                            'pe': 'Physical and Environmental Protection', 'pl': 'Planning',
                            'pm': 'Program Management', 'ps': 'Personnel Security',
                            'pt': 'PII Processing and Transparency', 'ra': 'Risk Assessment',
                            'sa': 'System and Services Acquisition', 'sc': 'System and Communications Protection',
                            'si': 'System and Information Integrity', 'sr': 'Supply Chain Risk Management'
                        }.get(family, family.upper())
                        
                        family_cell = ws_questions.cell(row=row, column=1)
                        family_cell.value = f"═══ {family.upper()} - {family_full_name} ═══"
                        family_cell.font = Font(bold=True, size=12, color='366092')
                        family_cell.alignment = Alignment(horizontal='left', vertical='center')
                        ws_questions.merge_cells(f'A{row}:F{row}')
                        for col in range(1, 7):
                            ws_questions.cell(row=row, column=col).fill = PatternFill(
                                start_color='E8EAF6', end_color='E8EAF6', fill_type='solid'
                            )
                        row += 1
                    
                    family_fill = PatternFill(
                        start_color=family_colors.get(family, 'FFFFFF'),
                        end_color=family_colors.get(family, 'FFFFFF'),
                        fill_type='solid'
                    )
                    
                    # Question type colors
                    question_type_colors = {
                        'implementation': 'E3F2FD',  # Light Blue
                        'evidence': 'FFF9C4',        # Light Yellow
                        'policy': 'F3E5F5',          # Light Purple
                        'procedure': 'E8F5E9',       # Light Green
                        'technical': 'FFE0B2',       # Light Orange
                        'administrative': 'FCE4EC',  # Light Pink
                        'physical': 'E0F2F1'         # Light Teal
                    }
                    
                    for q_num, q in enumerate(sorted_questions, 1):
                        q_type = q['question_type'].lower()
                        q_type_fill = PatternFill(
                            start_color=question_type_colors.get(q_type, 'FFFFFF'),
                            end_color=question_type_colors.get(q_type, 'FFFFFF'),
                            fill_type='solid'
                        )
                        
                        # Control ID
                        cell = ws_questions.cell(row=row, column=1)
                        cell.value = control_id
                        cell.font = Font(bold=True)
                        cell.fill = family_fill
                        cell.border = thin_border
                        cell.alignment = Alignment(horizontal='center', vertical='center')
                        
                        # Family
                        cell = ws_questions.cell(row=row, column=2)
                        cell.value = family.upper()
                        cell.fill = family_fill
                        cell.border = thin_border
                        cell.alignment = Alignment(horizontal='center', vertical='center')
                        
                        # Question Number
                        cell = ws_questions.cell(row=row, column=3)
                        cell.value = q_num
                        cell.fill = family_fill
                        cell.border = thin_border
                        cell.alignment = Alignment(horizontal='center', vertical='center')
                        
                        # Question Text
                        cell = ws_questions.cell(row=row, column=4)
                        cell.value = q['question_text']
                        cell.fill = family_fill
                        cell.border = thin_border
                        cell.alignment = Alignment(wrap_text=True, vertical='top')
                        
                        # Question Type (with color coding)
                        cell = ws_questions.cell(row=row, column=5)
                        cell.value = q['question_type'].upper()
                        cell.fill = q_type_fill
                        cell.border = thin_border
                        cell.alignment = Alignment(horizontal='center', vertical='center')
                        cell.font = Font(bold=True)
                        
                        # Response (blank)
                        cell = ws_questions.cell(row=row, column=6)
                        cell.value = ''
                        cell.fill = PatternFill(start_color='FFFACD', end_color='FFFACD', fill_type='solid')  # Light yellow for input
                        cell.border = thin_border
                        cell.alignment = Alignment(wrap_text=True, vertical='top')
                        
                        row += 1
            
            # Set column widths
            ws_questions.column_dimensions['A'].width = 15
            ws_questions.column_dimensions['B'].width = 10
            ws_questions.column_dimensions['C'].width = 10
            ws_questions.column_dimensions['D'].width = 60
            ws_questions.column_dimensions['E'].width = 20
            ws_questions.column_dimensions['F'].width = 60
            
            # Freeze header row
            ws_questions.freeze_panes = 'A2'
            
            # Save to bytes
            output = BytesIO()
            wb.save(output)
            output.seek(0)
            
            return output.getvalue(), 200, {
                'Content-Type': 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
                'Content-Disposition': 'attachment; filename=compliance-questionnaire.xlsx'
            }
        except Exception as e:
            return jsonify({'error': f'Excel export failed: {str(e)}'}), 500
    elif export_format == 'pdf':
        try:
            # Use simple PDF export (pure Python, no C extensions)
            from compliance_discovery.pdf_export_simple import generate_simple_pdf
            
            pdf_bytes = generate_simple_pdf(template_data)
            
            return pdf_bytes, 200, {
                'Content-Type': 'application/pdf',
                'Content-Disposition': 'attachment; filename=compliance-questionnaire.pdf'
            }
        except Exception as e:
            import traceback
            error_detail = traceback.format_exc()
            print(f"PDF export error: {error_detail}")
            return jsonify({
                'error': 'PDF export failed',
                'detail': str(e),
                'suggestion': 'Please use Excel export instead, which provides the same information.'
            }), 503
    else:
        return jsonify({'error': f'Unsupported format: {export_format}'}), 400


@app.route('/api/sessions', methods=['GET'])
def list_sessions():
    """List all sessions."""
    sessions = SessionModel.query.order_by(SessionModel.created_at.desc()).all()
    return jsonify({
        'sessions': [s.to_dict() for s in sessions],
        'total': len(sessions)
    })


@app.route('/api/ai/chat', methods=['POST'])
def ai_chat():
    """AI assistant chat endpoint powered by Amazon Bedrock.
    
    Request body:
        message: User's message/request
        engagement_context: Dict with config, schedule, evidence data
    """
    data = request.get_json()
    if not data or not data.get('message'):
        return jsonify({'error': 'Message is required'}), 400
    
    user_message = data['message']
    engagement_context = data.get('engagement_context', {})
    
    try:
        response_text = invoke_assistant(user_message, engagement_context)
        return jsonify({
            'response': response_text,
            'model': 'Claude 3.5 Haiku (Amazon Bedrock)',
        })
    except Exception as e:
        return jsonify({'error': f'AI assistant error: {str(e)}'}), 500


if __name__ == '__main__':
    print("Starting Compliance Discovery API Server...")
    print("Initializing database...")
    with app.app_context():
        db.create_all()
    print("Initializing data...")
    initialize_data('nist-800-53')
    initialize_data('nist-csf')
    print("Server ready!")
    app.run(debug=True, port=5001)
