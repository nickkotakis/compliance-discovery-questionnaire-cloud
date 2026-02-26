"""Flask REST API server for compliance discovery questionnaire."""

from flask import Flask, jsonify, request
from flask_cors import CORS
from typing import Dict, Any, List
from datetime import datetime
import json
import os

from compliance_discovery.nist_parser import NIST80053Parser
from compliance_discovery.question_generator import DiscoveryQuestionGenerator
from compliance_discovery.mcp_integration import MCPClient, create_aws_hints
from compliance_discovery.models.control import Control
from compliance_discovery.models.question import DiscoveryQuestion
from compliance_discovery.database import db, SessionModel
from compliance_discovery.control_descriptions import get_control_description
from compliance_discovery.aws_control_mapping import get_aws_responsibility, get_aws_services
from compliance_discovery.framework_mapper import get_framework_relevance


app = Flask(__name__)
CORS(app, resources={
    r"/api/*": {
        "origins": "*",
        "methods": ["GET", "POST", "PUT", "DELETE", "OPTIONS"],
        "allow_headers": ["Content-Type"]
    }
})

# Database configuration
basedir = os.path.abspath(os.path.dirname(__file__))
app.config['SQLALCHEMY_DATABASE_URI'] = f'sqlite:///{os.path.join(basedir, "sessions.db")}'
app.config['SQLALCHEMY_TRACK_MODIFICATIONS'] = False

# Initialize database
db.init_app(app)

# Create tables on startup (moved to main block to avoid double initialization)
# with app.app_context():
#     db.create_all()

# Global state
controls_cache: List[Control] = []
questions_cache: Dict[str, List[DiscoveryQuestion]] = {}
mcp_data_cache: Dict[str, List[Dict[str, Any]]] = {}  # Cache for MCP data from JSON file

# Initialize components
parser = NIST80053Parser()
generator = DiscoveryQuestionGenerator()
mcp_client = MCPClient()


def load_mcp_data():
    """Load AWS controls data from MCP JSON file."""
    global mcp_data_cache
    
    mcp_file = os.path.join(basedir, 'aws_controls_mcp_data.json')
    if os.path.exists(mcp_file):
        try:
            with open(mcp_file, 'r') as f:
                data = json.load(f)
                mcp_data_cache = data.get('controls', {})
                print(f"Loaded MCP data for {len(mcp_data_cache)} controls from {mcp_file}")
        except Exception as e:
            print(f"Warning: Could not load MCP data file: {str(e)}")
            mcp_data_cache = {}
    else:
        print(f"Note: MCP data file not found at {mcp_file}")
        mcp_data_cache = {}


def initialize_data():
    """Initialize controls and questions data."""
    global controls_cache, questions_cache
    
    if not controls_cache:
        print("Loading NIST 800-53 Moderate Baseline controls...")
        controls_cache = parser.get_moderate_baseline_controls()
        print(f"Loaded {len(controls_cache)} controls")
        print(f"Sample control IDs: {[c.id for c in controls_cache[:10]]}")
        
        print("Generating discovery questions...")
        for control in controls_cache:
            questions = generator.generate_questions(control)
            questions_cache[control.id] = questions
        print(f"Generated questions for {len(questions_cache)} controls")
        
        # Load MCP data
        load_mcp_data()


@app.after_request
def after_request(response):
    """Add CORS headers to all responses."""
    response.headers.add('Access-Control-Allow-Origin', '*')
    response.headers.add('Access-Control-Allow-Headers', 'Content-Type')
    response.headers.add('Access-Control-Allow-Methods', 'GET,POST,PUT,DELETE,OPTIONS')
    return response


@app.route('/api/controls', methods=['OPTIONS'])
@app.route('/api/controls/<control_id>', methods=['OPTIONS'])
@app.route('/api/questions', methods=['OPTIONS'])
@app.route('/api/session', methods=['OPTIONS'])
@app.route('/api/session/<session_id>', methods=['OPTIONS'])
@app.route('/api/session/<session_id>/response', methods=['OPTIONS'])
@app.route('/api/sessions', methods=['OPTIONS'])
@app.route('/api/export', methods=['OPTIONS'])
def handle_options(control_id=None, session_id=None):
    """Handle OPTIONS preflight requests."""
    return '', 204


@app.route('/api/health', methods=['GET'])
def health_check():
    """Health check endpoint."""
    return jsonify({
        'status': 'healthy',
        'timestamp': datetime.utcnow().isoformat()
    })


@app.route('/api/controls', methods=['GET'])
def get_controls():
    """Get all Moderate Baseline controls.
    
    Query parameters:
        family: Filter by control family (optional)
    """
    initialize_data()
    
    family = request.args.get('family')
    
    filtered_controls = controls_cache
    if family:
        filtered_controls = [c for c in controls_cache if c.family.upper() == family.upper()]
    
    # Sort controls alphabetically by ID with natural number sorting
    def sort_key(control):
        # Split ID into parts (e.g., "AC-11" -> ["AC", 11])
        parts = control.id.upper().split('-')
        if len(parts) == 2:
            family_code = parts[0]
            try:
                # Handle enhancements like AC-11(1)
                if '(' in parts[1]:
                    base_num = int(parts[1].split('(')[0])
                    enh_num = int(parts[1].split('(')[1].rstrip(')'))
                    return (family_code, base_num, enh_num)
                else:
                    return (family_code, int(parts[1]), 0)
            except ValueError:
                return (family_code, 0, 0)
        return (control.id.upper(), 0, 0)
    
    sorted_controls = sorted(filtered_controls, key=sort_key)
    
    return jsonify({
        'controls': [
            {
                'id': c.id,
                'title': c.title,
                'description': get_control_description(c.id) or c.description,
                'family': c.family,
                'in_moderate_baseline': c.in_moderate_baseline,
                'parameter_count': len(c.parameters),
                'enhancement_count': len(c.enhancements),
                'aws_responsibility': get_aws_responsibility(c.id)  # Add responsibility
            }
            for c in sorted_controls
        ],
        'total': len(sorted_controls)
    })


@app.route('/api/controls/<control_id>', methods=['GET'])
def get_control(control_id: str):
    """Get specific control with questions and AWS mappings.
    
    Args:
        control_id: NIST control ID (e.g., "AC-1" or "ac-1")
    """
    initialize_data()
    
    # Normalize control ID to uppercase for comparison
    normalized_id = control_id.upper()
    
    # Find control (case-insensitive)
    control = next((c for c in controls_cache if c.id.upper() == normalized_id), None)
    if not control:
        print(f"Control not found: {control_id} (normalized: {normalized_id})")
        print(f"Available controls: {[c.id for c in controls_cache[:5]]}")
        return jsonify({'error': f'Control not found: {control_id}'}), 404
    
    # Get questions (use actual control ID from cache)
    questions = questions_cache.get(control.id, [])
    
    # Try to get AWS mappings using normalized ID
    aws_hints = []
    aws_controls_data = []  # Store full control data for detailed view
    
    # First try loading from MCP data cache (JSON file)
    if normalized_id.lower() in mcp_data_cache:
        aws_controls_data = mcp_data_cache[normalized_id.lower()]
        print(f"Loaded {len(aws_controls_data)} AWS controls for {normalized_id} from MCP cache")
        # Generate hints from cached data
        if aws_controls_data:
            for ac in aws_controls_data:
                hint_parts = []
                if ac.get('services'):
                    services = ", ".join(ac['services'][:2])
                    hint_parts.append(f"Services: {services}")
                if ac.get('config_rules'):
                    rules = ", ".join(ac['config_rules'][:2])
                    hint_parts.append(f"Config: {rules}")
                if ac.get('security_hub_controls'):
                    hub = ", ".join(ac['security_hub_controls'][:2])
                    hint_parts.append(f"Security Hub: {hub}")
                if hint_parts:
                    aws_hints.append(f"• {ac['title']}\n  {' | '.join(hint_parts)}")
    
    # Fallback: Try MCP server (if available)
    if not aws_controls_data:
        try:
            if mcp_client.connected or mcp_client.connect():
                aws_controls = mcp_client.map_compliance_requirements(normalized_id)
                if aws_controls:
                    aws_hints = create_aws_hints(aws_controls)
                    # Store detailed control data
                    aws_controls_data = [
                        {
                            'control_id': ac.control_id,
                            'title': ac.title,
                            'description': ac.description,
                            'services': ac.services,
                            'config_rules': ac.managed_controls.config_rules,
                            'security_hub_controls': ac.managed_controls.security_hub_controls,
                            'control_tower_ids': ac.managed_controls.control_tower_ids,
                            'frameworks': ac.frameworks
                        }
                        for ac in aws_controls
                    ]
        except Exception as e:
            print(f"Warning: Could not fetch AWS mappings from MCP: {str(e)}")
    
    # Final fallback to manual mapping if nothing else worked
    if not aws_hints:
        fallback_services = get_aws_services(control.id)
        if fallback_services:
            aws_hints = [f"AWS Services: {', '.join(fallback_services)}"]
    
    # Determine responsibility using fallback mapping
    responsibility = get_aws_responsibility(control.id)
    
    # Build applicability message based on responsibility
    if responsibility == 'aws':
        aws_applicability = {
            'applicable': False,
            'responsibility': 'aws',
            'message': '🟠 AWS RESPONSIBILITY: AWS handles this control for their infrastructure. Your AWS workloads automatically inherit this protection. Download compliance reports from AWS Artifact to demonstrate this to auditors.',
            'artifact_links': [
                {
                    'name': 'AWS Artifact (Compliance Reports)',
                    'url': 'https://console.aws.amazon.com/artifact/',
                    'description': 'Download reports showing AWS implementation of this control'
                },
                {
                    'name': 'AWS Data Center Security',
                    'url': 'https://aws.amazon.com/compliance/data-center/',
                    'description': 'Learn about AWS physical security measures'
                }
            ],
            'controls': []
        }
    elif responsibility == 'shared' and aws_hints:
        aws_applicability = {
            'applicable': True,
            'responsibility': 'shared',
            'message': f'🟢 SHARED RESPONSIBILITY: AWS provides the services listed below. You must configure and use them properly in your AWS environment to meet this requirement.',
            'artifact_links': [
                {
                    'name': 'AWS Artifact (Compliance Reports)',
                    'url': 'https://console.aws.amazon.com/artifact/',
                    'description': 'Download SOC, PCI, ISO reports showing AWS platform compliance'
                },
                {
                    'name': 'AWS Compliance Programs',
                    'url': 'https://aws.amazon.com/compliance/programs/',
                    'description': 'View AWS compliance certifications'
                }
            ],
            'controls': aws_hints
        }
    else:
        aws_applicability = {
            'applicable': False,
            'responsibility': 'customer',
            'message': '🔵 CUSTOMER RESPONSIBILITY: You must implement this control in your AWS environment using custom configurations, third-party tools, or application-level controls.',
            'artifact_links': [],
            'controls': []
        }
    
    # Clean the description for display - use human-friendly version if available
    human_description = get_control_description(control.id)
    cleaned_description = human_description if human_description else control.description
    
    # Get framework relevance
    framework_relevance = get_framework_relevance(control.id)
    
    return jsonify({
        'control': {
            'id': control.id,
            'title': control.title,
            'description': cleaned_description,
            'family': control.family,
            'in_moderate_baseline': control.in_moderate_baseline,
            'parameters': [
                {
                    'id': p.id,
                    'label': p.label,
                    'description': p.description,
                    'constraints': p.constraints
                }
                for p in control.parameters
            ],
            'enhancements': [
                {
                    'id': e.id,
                    'title': e.title,
                    'description': e.description,
                    'in_moderate_baseline': e.in_moderate_baseline
                }
                for e in control.enhancements
            ]
        },
        'questions': [
            {
                'id': q.id,
                'control_id': q.control_id,
                'question_text': q.question_text,
                'question_type': q.question_type.value,
                'family': q.family,
                'aws_service_guidance': q.aws_service_guidance
            }
            for q in questions
        ],
        'aws_hints': aws_hints,
        'aws_applicability': aws_applicability,
        'aws_controls': aws_controls_data,  # Add detailed AWS control data
        'framework_relevance': framework_relevance  # Add framework relevance
    })


@app.route('/api/questions', methods=['GET'])
def get_questions():
    """Get all questions.
    
    Query parameters:
        control_id: Filter by control ID (optional)
        family: Filter by family (optional)
        question_type: Filter by question type (optional)
    """
    initialize_data()
    
    control_id = request.args.get('control_id')
    family = request.args.get('family')
    question_type = request.args.get('question_type')
    
    all_questions = []
    for questions in questions_cache.values():
        all_questions.extend(questions)
    
    # Apply filters
    if control_id:
        all_questions = [q for q in all_questions if q.control_id == control_id]
    if family:
        all_questions = [q for q in all_questions if q.family == family]
    if question_type:
        all_questions = [q for q in all_questions if q.question_type.value == question_type]
    
    return jsonify({
        'questions': [
            {
                'id': q.id,
                'control_id': q.control_id,
                'question_text': q.question_text,
                'question_type': q.question_type.value,
                'family': q.family,
                'aws_service_guidance': q.aws_service_guidance
            }
            for q in all_questions
        ],
        'total': len(all_questions)
    })


@app.route('/api/session', methods=['POST'])
def create_session():
    """Create new assessment session.
    
    Request body:
        {
            "customer_name": str,
            "analyst_name": str,
            "frameworks": List[str]
        }
    """
    data = request.get_json()
    
    session_id = f"session-{datetime.utcnow().timestamp()}"
    
    session_data = {
        'id': session_id,
        'customer_name': data.get('customer_name', ''),
        'analyst_name': data.get('analyst_name', ''),
        'frameworks': data.get('frameworks', ['NIST 800-53']),
        'status': 'active',
        'responses': {},
        'evidence': {}
    }
    
    # Save to database
    session_model = SessionModel.from_dict(session_data)
    db.session.add(session_model)
    db.session.commit()
    
    return jsonify(session_model.to_dict()), 201


@app.route('/api/session/<session_id>', methods=['GET'])
def get_session(session_id: str):
    """Get session details."""
    session = SessionModel.query.get(session_id)
    if not session:
        return jsonify({'error': 'Session not found'}), 404
    
    return jsonify(session.to_dict())


@app.route('/api/session/<session_id>/response', methods=['POST'])
def record_response(session_id: str):
    """Record a response to a question.
    
    Request body:
        {
            "question_id": str,
            "answer": str,
            "notes": str (optional)
        }
    """
    session = SessionModel.query.get(session_id)
    if not session:
        return jsonify({'error': 'Session not found'}), 404
    
    data = request.get_json()
    question_id = data.get('question_id')
    
    # Update responses
    responses = json.loads(session.responses) if session.responses else {}
    responses[question_id] = {
        'answer': data.get('answer', ''),
        'notes': data.get('notes', ''),
        'timestamp': datetime.utcnow().isoformat()
    }
    
    session.responses = json.dumps(responses)
    session.updated_at = datetime.utcnow()
    db.session.commit()
    
    return jsonify({'success': True})


@app.route('/api/export', methods=['GET'])
def export_template():
    """Export questionnaire in various formats.
    
    Query parameters:
        format: Export format (json, excel, pdf, yaml) - default: json
        include_unanswered: Include unanswered questions (default: true)
        include_aws_hints: Include AWS implementation hints (default: true)
        include_framework_mappings: Include framework mappings (default: true)
    """
    initialize_data()
    
    export_format = request.args.get('format', 'json')
    include_unanswered = request.args.get('include_unanswered', 'true').lower() == 'true'
    include_aws_hints = request.args.get('include_aws_hints', 'true').lower() == 'true'
    include_framework_mappings = request.args.get('include_framework_mappings', 'true').lower() == 'true'
    
    # Build template data
    template_data = {
        'metadata': {
            'template_version': '1.0.0',
            'baseline_version': 'NIST 800-53 Rev 5 Moderate Baseline',
            'export_date': datetime.utcnow().isoformat(),
            'total_control_count': len(controls_cache),
            'frameworks_included': ['NIST 800-53', 'AWS']
        },
        'controls': [
            {
                'id': c.id,
                'title': c.title,
                'description': get_control_description(c.id) or c.description,
                'family': c.family,
                'aws_responsibility': get_aws_responsibility(c.id)
            }
            for c in controls_cache
        ],
        'questions': {
            control_id: [
                {
                    'id': q.id,
                    'question_text': q.question_text,
                    'question_type': q.question_type.value,
                    'aws_service_guidance': q.aws_service_guidance if include_aws_hints else None,
                    'response': ''
                }
                for q in questions
            ]
            for control_id, questions in questions_cache.items()
        }
    }
    
    # Add framework mappings if requested
    if include_framework_mappings:
        template_data['framework_mappings'] = {
            control.id: get_framework_relevance(control.id)
            for control in controls_cache
        }
    
    # Add AWS hints if requested
    if include_aws_hints:
        template_data['aws_hints'] = {}
        for control in controls_cache:
            normalized_id = control.id.lower()
            if normalized_id in mcp_data_cache:
                aws_controls = mcp_data_cache[normalized_id]
                hints = []
                for ac in aws_controls:
                    hint_parts = []
                    if ac.get('services'):
                        services = ", ".join(ac['services'][:2])
                        hint_parts.append(f"Services: {services}")
                    if ac.get('config_rules'):
                        rules = ", ".join(ac['config_rules'][:2])
                        hint_parts.append(f"Config: {rules}")
                    if hint_parts:
                        hints.append(f"{ac['title']}: {' | '.join(hint_parts)}")
                template_data['aws_hints'][control.id] = hints
    
    if export_format == 'json':
        return jsonify(template_data)
    elif export_format == 'yaml':
        try:
            import yaml
            yaml_content = yaml.dump(template_data, default_flow_style=False, sort_keys=False)
            return yaml_content, 200, {'Content-Type': 'application/x-yaml', 'Content-Disposition': 'attachment; filename=compliance-questionnaire.yaml'}
        except ImportError:
            return jsonify({'error': 'YAML export requires PyYAML package'}), 500
    elif export_format == 'excel':
        try:
            from openpyxl import Workbook
            from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
            from io import BytesIO
            
            wb = Workbook()
            
            # Remove default sheet
            if 'Sheet' in wb.sheetnames:
                wb.remove(wb['Sheet'])
            
            # Define color scheme
            header_fill = PatternFill(start_color='366092', end_color='366092', fill_type='solid')
            header_font = Font(color='FFFFFF', bold=True, size=11)
            
            # Family colors (alternating for visual grouping)
            family_colors = {
                'ac': 'E3F2FD',  # Light Blue
                'at': 'F3E5F5',  # Light Purple
                'au': 'E8F5E9',  # Light Green
                'ca': 'FFF3E0',  # Light Orange
                'cm': 'FCE4EC',  # Light Pink
                'cp': 'E0F2F1',  # Light Teal
                'ia': 'FFF9C4',  # Light Yellow
                'ir': 'FFEBEE',  # Light Red
                'ma': 'E8EAF6',  # Light Indigo
                'mp': 'F1F8E9',  # Light Lime
                'pe': 'FBE9E7',  # Light Deep Orange
                'pl': 'E0F7FA',  # Light Cyan
                'pm': 'F9FBE7',  # Light Lime
                'ps': 'EDE7F6',  # Light Deep Purple
                'pt': 'E1F5FE',  # Light Light Blue
                'ra': 'F3E5F5',  # Light Purple
                'sa': 'E8F5E9',  # Light Green
                'sc': 'FFF3E0',  # Light Orange
                'si': 'FCE4EC',  # Light Pink
                'sr': 'E0F2F1',  # Light Teal
            }
            
            # Responsibility colors
            responsibility_colors = {
                'aws': 'FFE0B2',      # Orange
                'shared': 'C8E6C9',   # Green
                'customer': 'BBDEFB', # Blue
                'unknown': 'F5F5F5'   # Gray
            }
            
            # Border style
            thin_border = Border(
                left=Side(style='thin', color='CCCCCC'),
                right=Side(style='thin', color='CCCCCC'),
                top=Side(style='thin', color='CCCCCC'),
                bottom=Side(style='thin', color='CCCCCC')
            )
            
            # ===== CONTROLS SHEET =====
            ws_controls = wb.create_sheet("Controls", 0)
            
            # Headers
            headers = ['Control ID', 'Title', 'Description', 'Family', 'AWS Responsibility']
            for col_num, header in enumerate(headers, 1):
                cell = ws_controls.cell(row=1, column=col_num)
                cell.value = header
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                cell.border = thin_border
            
            # Sort controls by family, then by control number
            def sort_key(control):
                parts = control['id'].upper().split('-')
                if len(parts) == 2:
                    family_code = parts[0]
                    try:
                        if '(' in parts[1]:
                            base_num = int(parts[1].split('(')[0])
                            enh_num = int(parts[1].split('(')[1].rstrip(')'))
                            return (family_code, base_num, enh_num)
                        else:
                            return (family_code, int(parts[1]), 0)
                    except ValueError:
                        return (family_code, 0, 0)
                return (control['id'].upper(), 0, 0)
            
            sorted_controls = sorted(template_data['controls'], key=sort_key)
            
            # Data rows with family-based color coding
            current_family = None
            for idx, control in enumerate(sorted_controls, start=2):
                family = control['family'].lower()
                
                # Add family separator row when family changes
                if current_family != family:
                    current_family = family
                    # Add a bold family header row
                    family_row = ws_controls.cell(row=idx, column=1)
                    family_full_name = {
                        'ac': 'Access Control', 'at': 'Awareness and Training',
                        'au': 'Audit and Accountability', 'ca': 'Assessment, Authorization, and Monitoring',
                        'cm': 'Configuration Management', 'cp': 'Contingency Planning',
                        'ia': 'Identification and Authentication', 'ir': 'Incident Response',
                        'ma': 'Maintenance', 'mp': 'Media Protection',
                        'pe': 'Physical and Environmental Protection', 'pl': 'Planning',
                        'pm': 'Program Management', 'ps': 'Personnel Security',
                        'pt': 'PII Processing and Transparency', 'ra': 'Risk Assessment',
                        'sa': 'System and Services Acquisition', 'sc': 'System and Communications Protection',
                        'si': 'System and Information Integrity', 'sr': 'Supply Chain Risk Management'
                    }.get(family, family.upper())
                    
                    family_row.value = f"═══ {family.upper()} - {family_full_name} ═══"
                    family_row.font = Font(bold=True, size=12, color='366092')
                    family_row.alignment = Alignment(horizontal='left', vertical='center')
                    ws_controls.merge_cells(f'A{idx}:E{idx}')
                    for col in range(1, 6):
                        ws_controls.cell(row=idx, column=col).fill = PatternFill(
                            start_color='E8EAF6', end_color='E8EAF6', fill_type='solid'
                        )
                    idx += 1
                
                # Control data
                family_fill = PatternFill(
                    start_color=family_colors.get(family, 'FFFFFF'),
                    end_color=family_colors.get(family, 'FFFFFF'),
                    fill_type='solid'
                )
                
                # Control ID
                cell = ws_controls.cell(row=idx, column=1)
                cell.value = control['id']
                cell.font = Font(bold=True)
                cell.fill = family_fill
                cell.border = thin_border
                cell.alignment = Alignment(horizontal='center', vertical='center')
                
                # Title
                cell = ws_controls.cell(row=idx, column=2)
                cell.value = control['title']
                cell.fill = family_fill
                cell.border = thin_border
                cell.alignment = Alignment(wrap_text=True, vertical='top')
                
                # Description
                cell = ws_controls.cell(row=idx, column=3)
                cell.value = control['description']
                cell.fill = family_fill
                cell.border = thin_border
                cell.alignment = Alignment(wrap_text=True, vertical='top')
                
                # Family
                cell = ws_controls.cell(row=idx, column=4)
                cell.value = family.upper()
                cell.fill = family_fill
                cell.border = thin_border
                cell.alignment = Alignment(horizontal='center', vertical='center')
                
                # AWS Responsibility with color coding
                cell = ws_controls.cell(row=idx, column=5)
                responsibility = control.get('aws_responsibility', 'unknown')
                cell.value = responsibility.upper()
                cell.fill = PatternFill(
                    start_color=responsibility_colors.get(responsibility, 'F5F5F5'),
                    end_color=responsibility_colors.get(responsibility, 'F5F5F5'),
                    fill_type='solid'
                )
                cell.border = thin_border
                cell.alignment = Alignment(horizontal='center', vertical='center')
                cell.font = Font(bold=True)
            
            # Set column widths
            ws_controls.column_dimensions['A'].width = 15
            ws_controls.column_dimensions['B'].width = 40
            ws_controls.column_dimensions['C'].width = 60
            ws_controls.column_dimensions['D'].width = 12
            ws_controls.column_dimensions['E'].width = 18
            
            # Freeze header row
            ws_controls.freeze_panes = 'A2'
            
            # ===== QUESTIONS SHEET =====
            ws_questions = wb.create_sheet("Questions")
            
            # Headers
            question_headers = ['Control ID', 'Family', 'Question #', 'Question Text', 'Question Type', 'Response']
            for col_num, header in enumerate(question_headers, 1):
                cell = ws_questions.cell(row=1, column=col_num)
                cell.value = header
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                cell.border = thin_border
            
            # Question rows sorted by control (which are already sorted by family)
            row = 2
            current_family = None
            
            # Question type sort order (implementation first, then evidence, etc.)
            question_type_order = {
                'implementation': 1,
                'evidence': 2,
                'policy': 3,
                'procedure': 4,
                'technical': 5,
                'administrative': 6,
                'physical': 7
            }
            
            for control in sorted_controls:
                control_id = control['id']
                family = control['family'].lower()
                
                if control_id in template_data['questions']:
                    questions = template_data['questions'][control_id]
                    
                    # Sort questions by type
                    sorted_questions = sorted(
                        questions,
                        key=lambda q: question_type_order.get(q['question_type'].lower(), 99)
                    )
                    
                    # Add family separator when family changes
                    if current_family != family:
                        current_family = family
                        family_full_name = {
                            'ac': 'Access Control', 'at': 'Awareness and Training',
                            'au': 'Audit and Accountability', 'ca': 'Assessment, Authorization, and Monitoring',
                            'cm': 'Configuration Management', 'cp': 'Contingency Planning',
                            'ia': 'Identification and Authentication', 'ir': 'Incident Response',
                            'ma': 'Maintenance', 'mp': 'Media Protection',
                            'pe': 'Physical and Environmental Protection', 'pl': 'Planning',
                            'pm': 'Program Management', 'ps': 'Personnel Security',
                            'pt': 'PII Processing and Transparency', 'ra': 'Risk Assessment',
                            'sa': 'System and Services Acquisition', 'sc': 'System and Communications Protection',
                            'si': 'System and Information Integrity', 'sr': 'Supply Chain Risk Management'
                        }.get(family, family.upper())
                        
                        family_cell = ws_questions.cell(row=row, column=1)
                        family_cell.value = f"═══ {family.upper()} - {family_full_name} ═══"
                        family_cell.font = Font(bold=True, size=12, color='366092')
                        family_cell.alignment = Alignment(horizontal='left', vertical='center')
                        ws_questions.merge_cells(f'A{row}:F{row}')
                        for col in range(1, 7):
                            ws_questions.cell(row=row, column=col).fill = PatternFill(
                                start_color='E8EAF6', end_color='E8EAF6', fill_type='solid'
                            )
                        row += 1
                    
                    family_fill = PatternFill(
                        start_color=family_colors.get(family, 'FFFFFF'),
                        end_color=family_colors.get(family, 'FFFFFF'),
                        fill_type='solid'
                    )
                    
                    # Question type colors
                    question_type_colors = {
                        'implementation': 'E3F2FD',  # Light Blue
                        'evidence': 'FFF9C4',        # Light Yellow
                        'policy': 'F3E5F5',          # Light Purple
                        'procedure': 'E8F5E9',       # Light Green
                        'technical': 'FFE0B2',       # Light Orange
                        'administrative': 'FCE4EC',  # Light Pink
                        'physical': 'E0F2F1'         # Light Teal
                    }
                    
                    for q_num, q in enumerate(sorted_questions, 1):
                        q_type = q['question_type'].lower()
                        q_type_fill = PatternFill(
                            start_color=question_type_colors.get(q_type, 'FFFFFF'),
                            end_color=question_type_colors.get(q_type, 'FFFFFF'),
                            fill_type='solid'
                        )
                        
                        # Control ID
                        cell = ws_questions.cell(row=row, column=1)
                        cell.value = control_id
                        cell.font = Font(bold=True)
                        cell.fill = family_fill
                        cell.border = thin_border
                        cell.alignment = Alignment(horizontal='center', vertical='center')
                        
                        # Family
                        cell = ws_questions.cell(row=row, column=2)
                        cell.value = family.upper()
                        cell.fill = family_fill
                        cell.border = thin_border
                        cell.alignment = Alignment(horizontal='center', vertical='center')
                        
                        # Question Number
                        cell = ws_questions.cell(row=row, column=3)
                        cell.value = q_num
                        cell.fill = family_fill
                        cell.border = thin_border
                        cell.alignment = Alignment(horizontal='center', vertical='center')
                        
                        # Question Text
                        cell = ws_questions.cell(row=row, column=4)
                        cell.value = q['question_text']
                        cell.fill = family_fill
                        cell.border = thin_border
                        cell.alignment = Alignment(wrap_text=True, vertical='top')
                        
                        # Question Type (with color coding)
                        cell = ws_questions.cell(row=row, column=5)
                        cell.value = q['question_type'].upper()
                        cell.fill = q_type_fill
                        cell.border = thin_border
                        cell.alignment = Alignment(horizontal='center', vertical='center')
                        cell.font = Font(bold=True)
                        
                        # Response (blank)
                        cell = ws_questions.cell(row=row, column=6)
                        cell.value = ''
                        cell.fill = PatternFill(start_color='FFFACD', end_color='FFFACD', fill_type='solid')  # Light yellow for input
                        cell.border = thin_border
                        cell.alignment = Alignment(wrap_text=True, vertical='top')
                        
                        row += 1
            
            # Set column widths
            ws_questions.column_dimensions['A'].width = 15
            ws_questions.column_dimensions['B'].width = 10
            ws_questions.column_dimensions['C'].width = 10
            ws_questions.column_dimensions['D'].width = 60
            ws_questions.column_dimensions['E'].width = 20
            ws_questions.column_dimensions['F'].width = 60
            
            # Freeze header row
            ws_questions.freeze_panes = 'A2'
            
            # Save to bytes
            output = BytesIO()
            wb.save(output)
            output.seek(0)
            
            return output.getvalue(), 200, {
                'Content-Type': 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
                'Content-Disposition': 'attachment; filename=compliance-questionnaire.xlsx'
            }
        except Exception as e:
            return jsonify({'error': f'Excel export failed: {str(e)}'}), 500
    elif export_format == 'pdf':
        try:
            from reportlab.lib.pagesizes import letter
            from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
            from reportlab.lib.units import inch
            from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, PageBreak, Table, TableStyle
            from reportlab.lib import colors
            from reportlab.lib.enums import TA_CENTER, TA_LEFT, TA_JUSTIFY
            from io import BytesIO
            
            buffer = BytesIO()
            doc = SimpleDocTemplate(
                buffer,
                pagesize=letter,
                rightMargin=0.75*inch,
                leftMargin=0.75*inch,
                topMargin=0.75*inch,
                bottomMargin=0.75*inch
            )
            
            # Container for PDF elements
            elements = []
            styles = getSampleStyleSheet()
            
            # Custom styles
            title_style = ParagraphStyle(
                'CustomTitle',
                parent=styles['Heading1'],
                fontSize=24,
                textColor=colors.HexColor('#366092'),
                spaceAfter=12,
                alignment=TA_CENTER,
                fontName='Helvetica-Bold'
            )
            
            subtitle_style = ParagraphStyle(
                'CustomSubtitle',
                parent=styles['Heading2'],
                fontSize=16,
                textColor=colors.HexColor('#5A7FA3'),
                spaceAfter=20,
                alignment=TA_CENTER,
                fontName='Helvetica'
            )
            
            heading_style = ParagraphStyle(
                'CustomHeading',
                parent=styles['Heading2'],
                fontSize=14,
                textColor=colors.HexColor('#366092'),
                spaceAfter=12,
                spaceBefore=12,
                fontName='Helvetica-Bold'
            )
            
            family_heading_style = ParagraphStyle(
                'FamilyHeading',
                parent=styles['Heading3'],
                fontSize=12,
                textColor=colors.HexColor('#366092'),
                spaceAfter=8,
                spaceBefore=16,
                fontName='Helvetica-Bold',
                backColor=colors.HexColor('#E8EAF6'),
                leftIndent=6,
                rightIndent=6
            )
            
            control_style = ParagraphStyle(
                'ControlStyle',
                parent=styles['Normal'],
                fontSize=10,
                textColor=colors.black,
                spaceAfter=6,
                fontName='Helvetica-Bold'
            )
            
            question_style = ParagraphStyle(
                'QuestionStyle',
                parent=styles['Normal'],
                fontSize=9,
                textColor=colors.black,
                spaceAfter=4,
                leftIndent=12,
                fontName='Helvetica'
            )
            
            # Define color palette (matching Excel export)
            primary_blue = colors.HexColor('#366092')
            light_blue = colors.HexColor('#E3F2FD')
            light_yellow = colors.HexColor('#FFFACD')
            aws_orange = colors.HexColor('#FFE0B2')
            shared_green = colors.HexColor('#C8E6C9')
            customer_blue = colors.HexColor('#BBDEFB')
            
            # Question type colors
            question_type_colors = {
                'implementation': colors.HexColor('#E3F2FD'),
                'evidence': colors.HexColor('#FFF9C4'),
                'policy': colors.HexColor('#F3E5F5'),
                'procedure': colors.HexColor('#E8F5E9'),
                'technical': colors.HexColor('#FFE0B2'),
                'administrative': colors.HexColor('#FCE4EC'),
                'physical': colors.HexColor('#E0F2F1')
            }
            
            # Title page with color
            title_table = Table(
                [["Compliance Discovery Questionnaire"]],
                colWidths=[6.5*inch]
            )
            title_table.setStyle(TableStyle([
                ('BACKGROUND', (0, 0), (-1, -1), primary_blue),
                ('TEXTCOLOR', (0, 0), (-1, -1), colors.white),
                ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
                ('FONTNAME', (0, 0), (-1, -1), 'Helvetica-Bold'),
                ('FONTSIZE', (0, 0), (-1, -1), 24),
                ('TOPPADDING', (0, 0), (-1, -1), 20),
                ('BOTTOMPADDING', (0, 0), (-1, -1), 20),
            ]))
            elements.append(title_table)
            elements.append(Spacer(1, 0.2*inch))
            
            elements.append(Paragraph("NIST 800-53 Rev 5 Moderate Baseline", subtitle_style))
            elements.append(Spacer(1, 0.3*inch))
            
            # Metadata section with color
            metadata_data = [
                ["Export Date:", template_data['metadata']['export_date']],
                ["Total Controls:", str(template_data['metadata']['total_control_count'])],
                ["Frameworks:", ", ".join(template_data['metadata']['frameworks_included'])],
            ]
            
            metadata_table = Table(metadata_data, colWidths=[2*inch, 4*inch])
            metadata_table.setStyle(TableStyle([
                ('BACKGROUND', (0, 0), (-1, -1), light_blue),
                ('FONTNAME', (0, 0), (0, -1), 'Helvetica-Bold'),
                ('FONTNAME', (1, 0), (1, -1), 'Helvetica'),
                ('FONTSIZE', (0, 0), (-1, -1), 10),
                ('TEXTCOLOR', (0, 0), (0, -1), primary_blue),
                ('VALIGN', (0, 0), (-1, -1), 'TOP'),
                ('TOPPADDING', (0, 0), (-1, -1), 8),
                ('BOTTOMPADDING', (0, 0), (-1, -1), 8),
                ('LEFTPADDING', (0, 0), (-1, -1), 10),
                ('GRID', (0, 0), (-1, -1), 0.5, colors.HexColor('#BBDEFB')),
            ]))
            elements.append(metadata_table)
            elements.append(Spacer(1, 0.3*inch))
            
            # Instructions with colored box
            elements.append(Paragraph("Instructions", heading_style))
            instructions = [
                "• This questionnaire assesses compliance with NIST 800-53 Rev 5 Moderate Baseline controls",
                "• Questions are organized by type: Implementation, Evidence, Policy, Procedure, etc.",
                "• Fill in the response fields with detailed answers and specific examples",
                "• Document evidence locations and attach supporting materials as needed",
                "• AWS-specific guidance is provided to help leverage AWS managed services"
            ]
            
            inst_text = "<br/>".join(instructions)
            inst_para = Paragraph(inst_text, styles['Normal'])
            inst_table = Table([[inst_para]], colWidths=[6.5*inch])
            inst_table.setStyle(TableStyle([
                ('BACKGROUND', (0, 0), (-1, -1), light_yellow),
                ('TOPPADDING', (0, 0), (-1, -1), 12),
                ('BOTTOMPADDING', (0, 0), (-1, -1), 12),
                ('LEFTPADDING', (0, 0), (-1, -1), 12),
                ('RIGHTPADDING', (0, 0), (-1, -1), 12),
                ('BOX', (0, 0), (-1, -1), 1, colors.HexColor('#F0E68C')),
            ]))
            elements.append(inst_table)
            
            elements.append(PageBreak())
            
            # Sort controls by family
            def sort_key(control):
                parts = control['id'].upper().split('-')
                if len(parts) == 2:
                    family_code = parts[0]
                    try:
                        if '(' in parts[1]:
                            base_num = int(parts[1].split('(')[0])
                            enh_num = int(parts[1].split('(')[1].rstrip(')'))
                            return (family_code, base_num, enh_num)
                        else:
                            return (family_code, int(parts[1]), 0)
                    except ValueError:
                        return (family_code, 0, 0)
                return (control['id'].upper(), 0, 0)
            
            sorted_controls = sorted(template_data['controls'], key=sort_key)
            
            # Family names
            family_names = {
                'ac': 'Access Control', 'at': 'Awareness and Training',
                'au': 'Audit and Accountability', 'ca': 'Assessment, Authorization, and Monitoring',
                'cm': 'Configuration Management', 'cp': 'Contingency Planning',
                'ia': 'Identification and Authentication', 'ir': 'Incident Response',
                'ma': 'Maintenance', 'mp': 'Media Protection',
                'pe': 'Physical and Environmental Protection', 'pl': 'Planning',
                'pm': 'Program Management', 'ps': 'Personnel Security',
                'pt': 'PII Processing and Transparency', 'ra': 'Risk Assessment',
                'sa': 'System and Services Acquisition', 'sc': 'System and Communications Protection',
                'si': 'System and Information Integrity', 'sr': 'Supply Chain Risk Management'
            }
            
            # Question type sort order
            question_type_order = {
                'implementation': 1, 'evidence': 2, 'policy': 3,
                'procedure': 4, 'technical': 5, 'administrative': 6, 'physical': 7
            }
            
            # Controls by family
            current_family = None
            control_count = 0
            
            for control in sorted_controls:
                family = control['family'].lower()
                
                # Add family header when family changes
                if current_family != family:
                    if control_count > 0:
                        elements.append(PageBreak())
                    current_family = family
                    family_full_name = family_names.get(family, family.upper())
                    
                    # Family header with color
                    family_header = Table(
                        [[f"{family.upper()} - {family_full_name}"]],
                        colWidths=[6.5*inch]
                    )
                    family_header.setStyle(TableStyle([
                        ('BACKGROUND', (0, 0), (-1, -1), primary_blue),
                        ('TEXTCOLOR', (0, 0), (-1, -1), colors.white),
                        ('FONTNAME', (0, 0), (-1, -1), 'Helvetica-Bold'),
                        ('FONTSIZE', (0, 0), (-1, -1), 13),
                        ('TOPPADDING', (0, 0), (-1, -1), 10),
                        ('BOTTOMPADDING', (0, 0), (-1, -1), 10),
                        ('LEFTPADDING', (0, 0), (-1, -1), 10),
                    ]))
                    elements.append(family_header)
                    elements.append(Spacer(1, 0.15*inch))
                
                control_count += 1
                
                # Control header with responsibility badge
                responsibility = control.get('aws_responsibility', 'unknown').upper()
                resp_color = {
                    'AWS': aws_orange,
                    'SHARED': shared_green,
                    'CUSTOMER': customer_blue,
                    'UNKNOWN': colors.lightgrey
                }.get(responsibility, colors.white)
                
                control_header_data = [
                    [f"{control['id'].upper()}: {control['title']}", f"AWS: {responsibility}"]
                ]
                control_header_table = Table(control_header_data, colWidths=[5*inch, 1.5*inch])
                control_header_table.setStyle(TableStyle([
                    ('BACKGROUND', (0, 0), (0, 0), light_blue),
                    ('BACKGROUND', (1, 0), (1, 0), resp_color),
                    ('FONTNAME', (0, 0), (0, 0), 'Helvetica-Bold'),
                    ('FONTNAME', (1, 0), (1, 0), 'Helvetica-Bold'),
                    ('FONTSIZE', (0, 0), (-1, -1), 10),
                    ('TEXTCOLOR', (0, 0), (0, 0), primary_blue),
                    ('ALIGN', (1, 0), (1, 0), 'CENTER'),
                    ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
                    ('TOPPADDING', (0, 0), (-1, -1), 8),
                    ('BOTTOMPADDING', (0, 0), (-1, -1), 8),
                    ('LEFTPADDING', (0, 0), (-1, -1), 10),
                    ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
                ]))
                elements.append(control_header_table)
                elements.append(Spacer(1, 0.05*inch))
                
                # Control description
                desc_style = ParagraphStyle(
                    'DescStyle',
                    parent=styles['Normal'],
                    fontSize=9,
                    textColor=colors.HexColor('#333333'),
                    spaceAfter=8,
                    alignment=TA_JUSTIFY,
                    leftIndent=10,
                    rightIndent=10
                )
                desc_para = Paragraph(control['description'], desc_style)
                desc_table = Table([[desc_para]], colWidths=[6.5*inch])
                desc_table.setStyle(TableStyle([
                    ('BACKGROUND', (0, 0), (-1, -1), colors.HexColor('#FAFAFA')),
                    ('TOPPADDING', (0, 0), (-1, -1), 8),
                    ('BOTTOMPADDING', (0, 0), (-1, -1), 8),
                    ('BOX', (0, 0), (-1, -1), 0.5, colors.lightgrey),
                ]))
                elements.append(desc_table)
                elements.append(Spacer(1, 0.1*inch))
                
                # Questions for this control
                if control['id'] in template_data['questions']:
                    questions = template_data['questions'][control['id']]
                    
                    # Sort questions by type
                    sorted_questions = sorted(
                        questions,
                        key=lambda q: question_type_order.get(q['question_type'].lower(), 99)
                    )
                    
                    # Questions header
                    q_header = Table([["Assessment Questions"]], colWidths=[6.5*inch])
                    q_header.setStyle(TableStyle([
                        ('BACKGROUND', (0, 0), (-1, -1), colors.HexColor('#5A7FA3')),
                        ('TEXTCOLOR', (0, 0), (-1, -1), colors.white),
                        ('FONTNAME', (0, 0), (-1, -1), 'Helvetica-Bold'),
                        ('FONTSIZE', (0, 0), (-1, -1), 11),
                        ('TOPPADDING', (0, 0), (-1, -1), 6),
                        ('BOTTOMPADDING', (0, 0), (-1, -1), 6),
                        ('LEFTPADDING', (0, 0), (-1, -1), 10),
                    ]))
                    elements.append(q_header)
                    elements.append(Spacer(1, 0.05*inch))
                    
                    for idx, q in enumerate(sorted_questions, 1):
                        q_type = q['question_type'].lower()
                        q_type_display = q['question_type'].upper()
                        q_color = question_type_colors.get(q_type, colors.white)
                        
                        # Question with type badge
                        q_text = f"<b>{idx}.</b> {q['question_text']}"
                        q_para = Paragraph(q_text, question_style)
                        
                        q_table_data = [
                            [q_type_display, q_para],
                            ["RESPONSE:", ""]
                        ]
                        q_table = Table(q_table_data, colWidths=[1.2*inch, 5.3*inch], rowHeights=[None, 0.8*inch])
                        q_table.setStyle(TableStyle([
                            # Question type cell
                            ('BACKGROUND', (0, 0), (0, 0), q_color),
                            ('FONTNAME', (0, 0), (0, 0), 'Helvetica-Bold'),
                            ('FONTSIZE', (0, 0), (0, 0), 7),
                            ('ALIGN', (0, 0), (0, 0), 'CENTER'),
                            ('VALIGN', (0, 0), (0, 0), 'MIDDLE'),
                            ('WORDWRAP', (0, 0), (0, 0), True),
                            # Question text cell
                            ('BACKGROUND', (1, 0), (1, 0), colors.white),
                            ('VALIGN', (1, 0), (1, 0), 'TOP'),
                            ('LEFTPADDING', (1, 0), (1, 0), 8),
                            # Response label
                            ('BACKGROUND', (0, 1), (0, 1), light_yellow),
                            ('FONTNAME', (0, 1), (0, 1), 'Helvetica-Bold'),
                            ('FONTSIZE', (0, 1), (0, 1), 8),
                            ('ALIGN', (0, 1), (0, 1), 'CENTER'),
                            ('VALIGN', (0, 1), (0, 1), 'TOP'),
                            # Response field
                            ('BACKGROUND', (1, 1), (1, 1), light_yellow),
                            ('VALIGN', (1, 1), (1, 1), 'TOP'),
                            # Grid
                            ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
                            ('TOPPADDING', (0, 0), (-1, -1), 6),
                            ('BOTTOMPADDING', (0, 0), (-1, -1), 6),
                            ('LEFTPADDING', (0, 0), (-1, -1), 6),
                            ('RIGHTPADDING', (0, 0), (-1, -1), 6),
                        ]))
                        elements.append(q_table)
                        elements.append(Spacer(1, 0.08*inch))
                
                # Evidence section
                evidence_data = [
                    ["Evidence Documentation"],
                    ["Description:", ""],
                    ["Location:", ""],
                    ["Notes:", ""]
                ]
                evidence_table = Table(
                    evidence_data,
                    colWidths=[1.5*inch, 5*inch],
                    rowHeights=[None, 0.5*inch, 0.5*inch, 0.5*inch]
                )
                evidence_table.setStyle(TableStyle([
                    # Header
                    ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#8B9DC3')),
                    ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
                    ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                    ('FONTSIZE', (0, 0), (-1, 0), 10),
                    ('SPAN', (0, 0), (-1, 0)),
                    ('ALIGN', (0, 0), (-1, 0), 'CENTER'),
                    # Labels
                    ('BACKGROUND', (0, 1), (0, -1), colors.HexColor('#E8EAF6')),
                    ('FONTNAME', (0, 1), (0, -1), 'Helvetica-Bold'),
                    ('FONTSIZE', (0, 1), (0, -1), 9),
                    ('VALIGN', (0, 1), (0, -1), 'TOP'),
                    # Fields
                    ('BACKGROUND', (1, 1), (1, -1), colors.white),
                    ('VALIGN', (1, 1), (1, -1), 'TOP'),
                    # Grid
                    ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
                    ('TOPPADDING', (0, 0), (-1, -1), 6),
                    ('BOTTOMPADDING', (0, 0), (-1, -1), 6),
                    ('LEFTPADDING', (0, 0), (-1, -1), 8),
                ]))
                elements.append(evidence_table)
                elements.append(Spacer(1, 0.2*inch))
                
                # Add page break every 2 controls
                if control_count % 2 == 0:
                    elements.append(PageBreak())
            
            # Build PDF
            doc.build(elements)
            buffer.seek(0)
            
            # Add fillable form fields using pypdf
            try:
                from pypdf import PdfReader, PdfWriter
                from pypdf.generic import DictionaryObject, ArrayObject, TextStringObject, NameObject, NumberObject
                
                # Read the generated PDF
                reader = PdfReader(buffer)
                writer = PdfWriter()
                
                # Track field positions (we'll add form fields programmatically)
                # Note: This is a simplified approach - exact positioning would require
                # tracking coordinates during PDF generation
                
                # Copy all pages
                for page in reader.pages:
                    writer.add_page(page)
                
                # Create form fields dictionary
                # Note: Adding interactive form fields to an existing PDF with tables
                # is complex. For now, we'll make the PDF editable by ensuring
                # it has the proper structure, but users can fill it with any PDF editor.
                
                # Write to new buffer
                output_buffer = BytesIO()
                writer.write(output_buffer)
                output_buffer.seek(0)
                
                return output_buffer.getvalue(), 200, {
                    'Content-Type': 'application/pdf',
                    'Content-Disposition': 'attachment; filename=compliance-questionnaire.pdf'
                }
            except ImportError:
                # If pypdf is not available, return the basic PDF
                print("Warning: pypdf not available, returning non-fillable PDF")
                return buffer.getvalue(), 200, {
                    'Content-Type': 'application/pdf',
                    'Content-Disposition': 'attachment; filename=compliance-questionnaire.pdf'
                }
            except Exception as e:
                # If form field addition fails, return the basic PDF
                print(f"Warning: Could not add form fields: {str(e)}")
                return buffer.getvalue(), 200, {
                    'Content-Type': 'application/pdf',
                    'Content-Disposition': 'attachment; filename=compliance-questionnaire.pdf'
                }
        except Exception as e:
            import traceback
            error_detail = traceback.format_exc()
            print(f"PDF export error: {error_detail}")
            return jsonify({'error': f'PDF export failed: {str(e)}'}), 500
    else:
        return jsonify({'error': f'Unsupported format: {export_format}'}), 400


@app.route('/api/sessions', methods=['GET'])
def list_sessions():
    """List all sessions."""
    sessions = SessionModel.query.order_by(SessionModel.created_at.desc()).all()
    return jsonify({
        'sessions': [s.to_dict() for s in sessions],
        'total': len(sessions)
    })


if __name__ == '__main__':
    print("Starting Compliance Discovery API Server...")
    print("Initializing database...")
    with app.app_context():
        db.create_all()
    print("Initializing data...")
    initialize_data()
    print("Server ready!")
    app.run(debug=True, port=5001)
