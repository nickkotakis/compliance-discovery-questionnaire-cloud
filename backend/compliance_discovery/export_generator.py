"""Export generator for session reports and blank questionnaire templates."""
from __future__ import annotations

from typing import Dict, List, Optional, TYPE_CHECKING
from datetime import datetime

if TYPE_CHECKING:
    from openpyxl import Workbook

from .models import (
    Session,
    BlankQuestionnaireTemplate,
    TemplateMetadata,
    Control,
    DiscoveryQuestion,
    FrameworkMappings,
)
from .exceptions import (
    TemplateNotReadyError,
    TemplateGenerationError,
    ExcelGenerationError,
    PDFGenerationError,
)


# Template version constant
TEMPLATE_VERSION = "1.0.0"


class ExportGenerator:
    """Generates session reports and blank questionnaire templates.
    
    Attributes:
        controls: Loaded Moderate Baseline controls
        questions: Generated discovery questions organized by control_id
        framework_mappings: Framework mappings organized by control_id
        mcp_available: Whether MCP server is available for AWS mappings
    """
    
    def __init__(self):
        """Initialize the export generator."""
        self.controls: Optional[List[Control]] = None
        self.questions: Optional[Dict[str, List[DiscoveryQuestion]]] = None
        self.framework_mappings: Optional[Dict[str, FrameworkMappings]] = None
        self.mcp_available: bool = False
    
    def set_controls(self, controls: List[Control]) -> None:
        """Set the loaded controls.
        
        Args:
            controls: List of Moderate Baseline controls
        """
        self.controls = controls
    
    def set_questions(self, questions: Dict[str, List[DiscoveryQuestion]]) -> None:
        """Set the generated questions.
        
        Args:
            questions: Discovery questions organized by control_id
        """
        self.questions = questions
    
    def set_framework_mappings(self, mappings: Dict[str, FrameworkMappings]) -> None:
        """Set the framework mappings.
        
        Args:
            mappings: Framework mappings organized by control_id
        """
        self.framework_mappings = mappings
    
    def set_mcp_available(self, available: bool) -> None:
        """Set whether MCP server is available.
        
        Args:
            available: True if MCP server is available
        """
        self.mcp_available = available
    
    def is_ready_for_template_export(self) -> bool:
        """Check if all required data is loaded for template export.
        
        Returns:
            True if all data is ready
            
        Raises:
            TemplateNotReadyError: If required data is not yet loaded
        """
        missing_data = []
        
        if self.controls is None or len(self.controls) == 0:
            missing_data.append("controls not fully loaded from NIST")
        
        if self.questions is None or len(self.questions) == 0:
            missing_data.append("questions not generated")
        
        if self.framework_mappings is None or len(self.framework_mappings) == 0:
            missing_data.append("framework mappings not available")
        
        if missing_data:
            raise TemplateNotReadyError(
                f"Template export not ready. Missing: {', '.join(missing_data)}"
            )
        
        return True
    
    def get_template_version(self) -> str:
        """Return current template version string.
        
        Returns:
            Template version in semantic versioning format (e.g., "1.0.0")
        """
        return TEMPLATE_VERSION
    
    # Session export methods
    def export_json(self, session: Session) -> str:
        """Export session as JSON (machine-readable).
        
        Includes all responses, evidence documentation, and metadata.
        
        Args:
            session: Session to export
            
        Returns:
            JSON string representation of the session
        """
        # TODO: Implement JSON export
        raise NotImplementedError("JSON export not yet implemented")
    
    def export_markdown(self, session: Session) -> str:
        """Export session as Markdown (human-readable).
        
        Includes all responses, evidence documentation organized by control,
        and framework coverage analysis.
        
        Args:
            session: Session to export
            
        Returns:
            Markdown string representation of the session
        """
        # TODO: Implement Markdown export
        raise NotImplementedError("Markdown export not yet implemented")
    
    def export_csv(self, session: Session) -> str:
        """Export session responses and evidence as CSV.
        
        Includes separate sections for responses and evidence entries.
        
        Args:
            session: Session to export
            
        Returns:
            CSV string representation of the session
        """
        # TODO: Implement CSV export
        raise NotImplementedError("CSV export not yet implemented")
    
    # Template export methods (to be implemented in later phases)
    def export_blank_template_excel(self, template: BlankQuestionnaireTemplate) -> bytes:
        """Generate Excel workbook with multiple sheets for blank template.
        
        Creates separate worksheets for:
        - Controls (all Moderate Baseline controls with descriptions)
        - Questions (all discovery questions organized by control)
        - Framework Mappings (NIST CSF, GLBA, SOX, FFIEC mappings)
        - AWS Hints (AWS Config rules, Security Hub controls, Control Tower IDs)
        - Instructions (how to complete the template)
        - Metadata (template version, baseline version, export date)
        - Evidence (blank evidence fields for each control)
        
        Args:
            template: BlankQuestionnaireTemplate object with all data
            
        Returns:
            Excel workbook as bytes (.xlsx format)
            
        Raises:
            ExcelGenerationError: If Excel generation fails
        """
        try:
            from openpyxl import Workbook
            from openpyxl.styles import Font, Alignment
            from io import BytesIO
            from .excel_helpers import (
                apply_header_formatting,
                set_column_widths,
                freeze_header_row,
                enable_text_wrap,
            )
            
            wb = Workbook()
            
            # Remove default sheet
            if "Sheet" in wb.sheetnames:
                wb.remove(wb["Sheet"])
            
            # Create Instructions worksheet first
            self._create_instructions_worksheet(wb, template)
            
            # Create Metadata worksheet
            self._create_metadata_worksheet(wb, template)
            
            # Create Controls worksheet
            self._create_controls_worksheet(wb, template)
            
            # Create Questions worksheet
            self._create_questions_worksheet(wb, template)
            
            # Create Framework Mappings worksheet
            self._create_framework_mappings_worksheet(wb, template)
            
            # Create AWS Hints worksheet
            self._create_aws_hints_worksheet(wb, template)
            
            # Create Evidence worksheet
            self._create_evidence_worksheet(wb, template)
            
            # Save to bytes
            output = BytesIO()
            wb.save(output)
            output.seek(0)
            return output.getvalue()
            
        except Exception as e:
            raise ExcelGenerationError(f"Failed to generate Excel template: {str(e)}")
    
    def _create_instructions_worksheet(self, wb: Workbook, template: BlankQuestionnaireTemplate):
        """Create Instructions worksheet."""
        from openpyxl.styles import Font, Alignment
        
        ws = wb.create_sheet("Instructions", 0)
        
        # Add title
        ws["A1"] = "Compliance Discovery Questionnaire - Instructions"
        ws["A1"].font = Font(bold=True, size=14)
        
        # Add instructions text
        instructions_lines = template.metadata.instructions.split("\n")
        for idx, line in enumerate(instructions_lines, start=3):
            ws[f"A{idx}"] = line
            ws[f"A{idx}"].alignment = Alignment(wrap_text=True, vertical="top")
        
        # Set column width
        ws.column_dimensions["A"].width = 120
    
    def _create_metadata_worksheet(self, wb: Workbook, template: BlankQuestionnaireTemplate):
        """Create Metadata worksheet."""
        from openpyxl.styles import Font
        from .excel_helpers import apply_header_formatting, set_column_widths
        
        ws = wb.create_sheet("Metadata")
        
        # Headers
        headers = ["Key", "Value"]
        apply_header_formatting(ws, headers)
        
        # Metadata rows
        metadata_items = [
            ("Template Version", template.metadata.template_version),
            ("Baseline Version", template.metadata.baseline_version),
            ("Export Date", template.metadata.export_date.strftime("%Y-%m-%d %H:%M:%S")),
            ("Total Control Count", str(template.metadata.total_control_count)),
            ("Frameworks Included", ", ".join(template.metadata.frameworks_included)),
        ]
        
        for idx, (key, value) in enumerate(metadata_items, start=2):
            ws[f"A{idx}"] = key
            ws[f"A{idx}"].font = Font(bold=True)
            ws[f"B{idx}"] = value
        
        set_column_widths(ws, {"A": 25, "B": 60})
    
    def _create_controls_worksheet(self, wb: Workbook, template: BlankQuestionnaireTemplate):
        """Create Controls worksheet."""
        from .excel_helpers import (
            apply_header_formatting,
            set_column_widths,
            freeze_header_row,
            enable_text_wrap,
        )
        
        ws = wb.create_sheet("Controls")
        
        # Headers
        headers = ["Control ID", "Title", "Description", "Family"]
        apply_header_formatting(ws, headers)
        
        # Control rows
        for idx, control in enumerate(template.controls, start=2):
            ws[f"A{idx}"] = control.id
            ws[f"B{idx}"] = control.title
            ws[f"C{idx}"] = control.description
            ws[f"D{idx}"] = control.family
        
        set_column_widths(ws, {"A": 15, "B": 40, "C": 60, "D": 15})
        freeze_header_row(ws)
        enable_text_wrap(ws, ["B", "C"])
    
    def _create_questions_worksheet(self, wb: Workbook, template: BlankQuestionnaireTemplate):
        """Create Questions worksheet."""
        from .excel_helpers import (
            apply_header_formatting,
            set_column_widths,
            freeze_header_row,
            enable_text_wrap,
        )
        
        ws = wb.create_sheet("Questions")
        
        # Headers
        headers = ["Question ID", "Control ID", "Question Text", "Question Type", "Response"]
        apply_header_formatting(ws, headers)
        
        # Question rows
        row = 2
        for control_id in sorted(template.questions.keys()):
            questions = template.questions[control_id]
            for question in questions:
                ws[f"A{row}"] = question.id
                ws[f"B{row}"] = question.control_id
                ws[f"C{row}"] = question.question_text
                ws[f"D{row}"] = question.question_type.value
                ws[f"E{row}"] = ""  # Blank response field
                row += 1
        
        set_column_widths(ws, {"A": 20, "B": 15, "C": 60, "D": 20, "E": 60})
        freeze_header_row(ws)
        enable_text_wrap(ws, ["C", "E"])
    
    def _create_framework_mappings_worksheet(self, wb: Workbook, template: BlankQuestionnaireTemplate):
        """Create Framework Mappings worksheet."""
        from .excel_helpers import (
            apply_header_formatting,
            set_column_widths,
            freeze_header_row,
            enable_text_wrap,
        )
        
        ws = wb.create_sheet("Framework Mappings")
        
        # Headers
        headers = ["Control ID", "Framework", "Requirement ID", "Description"]
        apply_header_formatting(ws, headers)
        
        # Mapping rows
        row = 2
        for control_id in sorted(template.framework_mappings.keys()):
            mappings = template.framework_mappings[control_id]
            
            # NIST CSF mappings
            for mapping in mappings.nist_csf:
                ws[f"A{row}"] = control_id
                ws[f"B{row}"] = "NIST CSF"
                ws[f"C{row}"] = mapping.subcategory
                ws[f"D{row}"] = f"{mapping.function} - {mapping.category}"
                row += 1
            
            # GLBA mappings
            for mapping in mappings.glba:
                ws[f"A{row}"] = control_id
                ws[f"B{row}"] = "GLBA"
                ws[f"C{row}"] = mapping.requirement_id
                ws[f"D{row}"] = mapping.description
                row += 1
            
            # SOX mappings
            for mapping in mappings.sox:
                ws[f"A{row}"] = control_id
                ws[f"B{row}"] = "SOX"
                ws[f"C{row}"] = f"Section {mapping.section}"
                ws[f"D{row}"] = f"{mapping.control_type} - {mapping.description}"
                row += 1
            
            # FFIEC mappings
            for mapping in mappings.ffiec:
                ws[f"A{row}"] = control_id
                ws[f"B{row}"] = "FFIEC"
                ws[f"C{row}"] = mapping.assessment_factor
                ws[f"D{row}"] = f"{mapping.domain} - {mapping.description}"
                row += 1
        
        set_column_widths(ws, {"A": 15, "B": 15, "C": 30, "D": 60})
        freeze_header_row(ws)
        enable_text_wrap(ws, ["D"])
    
    def _create_aws_hints_worksheet(self, wb: Workbook, template: BlankQuestionnaireTemplate):
        """Create AWS Hints worksheet."""
        from .excel_helpers import (
            apply_header_formatting,
            set_column_widths,
            freeze_header_row,
            enable_text_wrap,
        )
        
        ws = wb.create_sheet("AWS Hints")
        
        # Headers
        headers = ["Control ID", "AWS Control Hints"]
        apply_header_formatting(ws, headers)
        
        # Add note if MCP unavailable
        if not self.mcp_available:
            ws["A2"] = "NOTE"
            ws["B2"] = "AWS control hints are not available. MCP server connection required."
            ws["A2"].font = Font(bold=True)
        
        # AWS hints rows
        row = 3 if not self.mcp_available else 2
        for control_id in sorted(template.aws_hints.keys()):
            hints = template.aws_hints[control_id]
            if hints:
                for hint in hints:
                    ws[f"A{row}"] = control_id
                    ws[f"B{row}"] = hint
                    row += 1
            elif self.mcp_available:
                # Control has no AWS mappings
                ws[f"A{row}"] = control_id
                ws[f"B{row}"] = "No AWS control mappings available"
                row += 1
        
        set_column_widths(ws, {"A": 15, "B": 80})
        freeze_header_row(ws)
        enable_text_wrap(ws, ["B"])
    
    def _create_evidence_worksheet(self, wb: Workbook, template: BlankQuestionnaireTemplate):
        """Create Evidence worksheet."""
        from .excel_helpers import (
            apply_header_formatting,
            set_column_widths,
            freeze_header_row,
            enable_text_wrap,
        )
        
        ws = wb.create_sheet("Evidence")
        
        # Headers
        headers = ["Control ID", "Evidence Description", "Evidence Location", "Notes"]
        apply_header_formatting(ws, headers)
        
        # Evidence rows (blank fields for each control)
        for idx, control in enumerate(template.controls, start=2):
            ws[f"A{idx}"] = control.id
            ws[f"B{idx}"] = ""  # Blank evidence description
            ws[f"C{idx}"] = ""  # Blank evidence location
            ws[f"D{idx}"] = ""  # Blank notes
        
        set_column_widths(ws, {"A": 15, "B": 40, "C": 40, "D": 40})
        freeze_header_row(ws)
        enable_text_wrap(ws, ["B", "C", "D"])
    
    def export_blank_template_csv(self, template: BlankQuestionnaireTemplate) -> Dict[str, str]:
        """Generate CSV files for blank template.
        
        Creates separate CSV files for:
        - controls.csv (control_id, title, description, family)
        - questions.csv (question_id, control_id, question_text, question_type)
        - mappings.csv (control_id, framework, requirement_id, description)
        - aws_hints.csv (control_id, hint_text)
        - metadata.csv (key, value pairs)
        - evidence.csv (control_id, evidence_description, evidence_location, notes)
        
        Args:
            template: BlankQuestionnaireTemplate object with all data
            
        Returns:
            Dictionary mapping filename to CSV content string
            
        Raises:
            TemplateGenerationError: If CSV generation fails
        """
        try:
            import csv
            from io import StringIO
            
            csv_files = {}
            
            # Controls CSV
            csv_files["controls.csv"] = self._generate_controls_csv(template)
            
            # Questions CSV
            csv_files["questions.csv"] = self._generate_questions_csv(template)
            
            # Mappings CSV
            csv_files["mappings.csv"] = self._generate_mappings_csv(template)
            
            # AWS Hints CSV
            csv_files["aws_hints.csv"] = self._generate_aws_hints_csv(template)
            
            # Metadata CSV
            csv_files["metadata.csv"] = self._generate_metadata_csv(template)
            
            # Evidence CSV
            csv_files["evidence.csv"] = self._generate_evidence_csv(template)
            
            return csv_files
            
        except Exception as e:
            raise TemplateGenerationError(f"Failed to generate CSV templates: {str(e)}")
    
    def _generate_controls_csv(self, template: BlankQuestionnaireTemplate) -> str:
        """Generate controls.csv content."""
        import csv
        from io import StringIO
        
        output = StringIO()
        writer = csv.writer(output)
        
        # Header
        writer.writerow(["control_id", "title", "description", "family"])
        
        # Data rows
        for control in template.controls:
            writer.writerow([
                control.id,
                control.title,
                control.description,
                control.family,
            ])
        
        return output.getvalue()
    
    def _generate_questions_csv(self, template: BlankQuestionnaireTemplate) -> str:
        """Generate questions.csv content."""
        import csv
        from io import StringIO
        
        output = StringIO()
        writer = csv.writer(output)
        
        # Header
        writer.writerow(["question_id", "control_id", "question_text", "question_type", "response"])
        
        # Data rows
        for control_id in sorted(template.questions.keys()):
            questions = template.questions[control_id]
            for question in questions:
                writer.writerow([
                    question.id,
                    question.control_id,
                    question.question_text,
                    question.question_type.value,
                    "",  # Blank response field
                ])
        
        return output.getvalue()
    
    def _generate_mappings_csv(self, template: BlankQuestionnaireTemplate) -> str:
        """Generate mappings.csv content."""
        import csv
        from io import StringIO
        
        output = StringIO()
        writer = csv.writer(output)
        
        # Header
        writer.writerow(["control_id", "framework", "requirement_id", "description"])
        
        # Data rows
        for control_id in sorted(template.framework_mappings.keys()):
            mappings = template.framework_mappings[control_id]
            
            # NIST CSF mappings
            for mapping in mappings.nist_csf:
                writer.writerow([
                    control_id,
                    "NIST CSF",
                    mapping.subcategory,
                    f"{mapping.function} - {mapping.category}",
                ])
            
            # GLBA mappings
            for mapping in mappings.glba:
                writer.writerow([
                    control_id,
                    "GLBA",
                    mapping.requirement_id,
                    mapping.description,
                ])
            
            # SOX mappings
            for mapping in mappings.sox:
                writer.writerow([
                    control_id,
                    "SOX",
                    f"Section {mapping.section}",
                    f"{mapping.control_type} - {mapping.description}",
                ])
            
            # FFIEC mappings
            for mapping in mappings.ffiec:
                writer.writerow([
                    control_id,
                    "FFIEC",
                    mapping.assessment_factor,
                    f"{mapping.domain} - {mapping.description}",
                ])
        
        return output.getvalue()
    
    def _generate_aws_hints_csv(self, template: BlankQuestionnaireTemplate) -> str:
        """Generate aws_hints.csv content."""
        import csv
        from io import StringIO
        
        output = StringIO()
        writer = csv.writer(output)
        
        # Header
        writer.writerow(["control_id", "hint_text"])
        
        # Add note if MCP unavailable
        if not self.mcp_available:
            writer.writerow(["NOTE", "AWS control hints are not available. MCP server connection required."])
        
        # Data rows
        for control_id in sorted(template.aws_hints.keys()):
            hints = template.aws_hints[control_id]
            if hints:
                for hint in hints:
                    writer.writerow([control_id, hint])
            elif self.mcp_available:
                writer.writerow([control_id, "No AWS control mappings available"])
        
        return output.getvalue()
    
    def _generate_metadata_csv(self, template: BlankQuestionnaireTemplate) -> str:
        """Generate metadata.csv content."""
        import csv
        from io import StringIO
        
        output = StringIO()
        writer = csv.writer(output)
        
        # Header
        writer.writerow(["key", "value"])
        
        # Data rows
        writer.writerow(["template_version", template.metadata.template_version])
        writer.writerow(["baseline_version", template.metadata.baseline_version])
        writer.writerow(["export_date", template.metadata.export_date.strftime("%Y-%m-%d %H:%M:%S")])
        writer.writerow(["total_control_count", str(template.metadata.total_control_count)])
        writer.writerow(["frameworks_included", ", ".join(template.metadata.frameworks_included)])
        
        return output.getvalue()
    
    def _generate_evidence_csv(self, template: BlankQuestionnaireTemplate) -> str:
        """Generate evidence.csv content."""
        import csv
        from io import StringIO
        
        output = StringIO()
        writer = csv.writer(output)
        
        # Header
        writer.writerow(["control_id", "evidence_description", "evidence_location", "notes"])
        
        # Data rows (blank fields for each control)
        for control in template.controls:
            writer.writerow([
                control.id,
                "",  # Blank evidence description
                "",  # Blank evidence location
                "",  # Blank notes
            ])
        
        return output.getvalue()
    
    def export_blank_template_pdf(self, template: BlankQuestionnaireTemplate) -> bytes:
        """Generate PDF with formatted sections for blank template.
        
        Includes:
        - Table of contents
        - Metadata section
        - Instructions section
        - Controls organized by family with questions and blank response fields
        - Framework mappings section
        - AWS hints section
        
        Args:
            template: BlankQuestionnaireTemplate object with all data
            
        Returns:
            PDF document as bytes
            
        Raises:
            PDFGenerationError: If PDF generation fails
        """
        try:
            from reportlab.lib.pagesizes import letter
            from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
            from reportlab.lib.units import inch
            from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, PageBreak, Table, TableStyle
            from reportlab.lib import colors
            from io import BytesIO
            
            buffer = BytesIO()
            doc = SimpleDocTemplate(buffer, pagesize=letter,
                                   rightMargin=0.75*inch, leftMargin=0.75*inch,
                                   topMargin=0.75*inch, bottomMargin=0.75*inch)
            
            # Container for PDF elements
            elements = []
            styles = getSampleStyleSheet()
            
            # Custom styles
            title_style = ParagraphStyle(
                'CustomTitle',
                parent=styles['Heading1'],
                fontSize=18,
                textColor=colors.HexColor('#366092'),
                spaceAfter=30,
            )
            
            heading_style = ParagraphStyle(
                'CustomHeading',
                parent=styles['Heading2'],
                fontSize=14,
                textColor=colors.HexColor('#366092'),
                spaceAfter=12,
            )
            
            # Title page
            elements.append(Paragraph("Compliance Discovery Questionnaire", title_style))
            elements.append(Paragraph("NIST 800-53 Rev 5 Moderate Baseline", styles['Heading2']))
            elements.append(Spacer(1, 0.5*inch))
            
            # Metadata section
            elements.append(Paragraph("Template Metadata", heading_style))
            metadata_data = [
                ["Template Version:", template.metadata.template_version],
                ["Baseline Version:", template.metadata.baseline_version],
                ["Export Date:", template.metadata.export_date.strftime("%Y-%m-%d %H:%M:%S")],
                ["Total Controls:", str(template.metadata.total_control_count)],
                ["Frameworks:", ", ".join(template.metadata.frameworks_included)],
            ]
            metadata_table = Table(metadata_data, colWidths=[2*inch, 4*inch])
            metadata_table.setStyle(TableStyle([
                ('FONTNAME', (0, 0), (0, -1), 'Helvetica-Bold'),
                ('FONTNAME', (1, 0), (1, -1), 'Helvetica'),
                ('FONTSIZE', (0, 0), (-1, -1), 10),
                ('VALIGN', (0, 0), (-1, -1), 'TOP'),
                ('BOTTOMPADDING', (0, 0), (-1, -1), 6),
            ]))
            elements.append(metadata_table)
            elements.append(PageBreak())
            
            # Instructions section
            elements.append(Paragraph("Instructions", heading_style))
            for line in template.metadata.instructions.split("\n"):
                if line.strip():
                    if line.startswith("#"):
                        # Heading
                        elements.append(Paragraph(line.replace("#", "").strip(), styles['Heading3']))
                    else:
                        elements.append(Paragraph(line, styles['Normal']))
                else:
                    elements.append(Spacer(1, 0.1*inch))
            elements.append(PageBreak())
            
            # Controls by family
            elements.append(Paragraph("Controls by Family", heading_style))
            
            # Group controls by family
            controls_by_family = {}
            for control in template.controls:
                if control.family not in controls_by_family:
                    controls_by_family[control.family] = []
                controls_by_family[control.family].append(control)
            
            for family in sorted(controls_by_family.keys()):
                elements.append(Paragraph(f"Family: {family}", styles['Heading3']))
                
                for control in controls_by_family[family]:
                    # Control header
                    elements.append(Paragraph(f"<b>{control.id}: {control.title}</b>", styles['Normal']))
                    elements.append(Paragraph(control.description, styles['Normal']))
                    elements.append(Spacer(1, 0.1*inch))
                    
                    # Questions for this control
                    if control.id in template.questions:
                        elements.append(Paragraph("<b>Assessment Questions:</b>", styles['Normal']))
                        for question in template.questions[control.id]:
                            elements.append(Paragraph(f"• {question.question_text}", styles['Normal']))
                            elements.append(Paragraph("<i>Response: _______________</i>", styles['Normal']))
                            elements.append(Spacer(1, 0.05*inch))
                    
                    # Evidence fields
                    elements.append(Paragraph("<b>Evidence:</b>", styles['Normal']))
                    elements.append(Paragraph("Description: _______________", styles['Normal']))
                    elements.append(Paragraph("Location: _______________", styles['Normal']))
                    elements.append(Paragraph("Notes: _______________", styles['Normal']))
                    elements.append(Spacer(1, 0.2*inch))
                
                elements.append(PageBreak())
            
            # Build PDF
            doc.build(elements)
            buffer.seek(0)
            return buffer.getvalue()
            
        except Exception as e:
            raise PDFGenerationError(f"Failed to generate PDF template: {str(e)}")
    
    def export_blank_template_json(self, template: BlankQuestionnaireTemplate) -> str:
        """Generate JSON with structured schema for blank template.
        
        Args:
            template: BlankQuestionnaireTemplate object with all data
            
        Returns:
            JSON string with structured template data
            
        Raises:
            TemplateGenerationError: If JSON generation fails
        """
        try:
            import json
            from dataclasses import asdict
            
            # Convert template to dictionary
            template_dict = {
                "metadata": {
                    "template_version": template.metadata.template_version,
                    "baseline_version": template.metadata.baseline_version,
                    "export_date": template.metadata.export_date.isoformat(),
                    "total_control_count": template.metadata.total_control_count,
                    "frameworks_included": template.metadata.frameworks_included,
                    "instructions": template.metadata.instructions,
                },
                "controls": [
                    {
                        "id": control.id,
                        "title": control.title,
                        "description": control.description,
                        "family": control.family,
                        "in_moderate_baseline": control.in_moderate_baseline,
                    }
                    for control in template.controls
                ],
                "questions": {
                    control_id: [
                        {
                            "id": q.id,
                            "control_id": q.control_id,
                            "question_text": q.question_text,
                            "question_type": q.question_type.value,
                            "family": q.family,
                            "response": "",  # Blank response field
                        }
                        for q in questions
                    ]
                    for control_id, questions in template.questions.items()
                },
                "framework_mappings": {
                    control_id: {
                        "nist_csf": [
                            {
                                "function": m.function,
                                "category": m.category,
                                "subcategory": m.subcategory,
                            }
                            for m in mappings.nist_csf
                        ],
                        "glba": [
                            {
                                "requirement_type": m.requirement_type,
                                "requirement_id": m.requirement_id,
                                "description": m.description,
                            }
                            for m in mappings.glba
                        ],
                        "sox": [
                            {
                                "section": m.section,
                                "control_type": m.control_type,
                                "description": m.description,
                            }
                            for m in mappings.sox
                        ],
                        "ffiec": [
                            {
                                "domain": m.domain,
                                "assessment_factor": m.assessment_factor,
                                "description": m.description,
                            }
                            for m in mappings.ffiec
                        ],
                    }
                    for control_id, mappings in template.framework_mappings.items()
                },
                "aws_hints": template.aws_hints,
                "evidence": {
                    control.id: {
                        "description": "",
                        "location": "",
                        "notes": "",
                    }
                    for control in template.controls
                },
            }
            
            return json.dumps(template_dict, indent=2)
            
        except Exception as e:
            raise TemplateGenerationError(f"Failed to generate JSON template: {str(e)}")
    
    def export_blank_template_yaml(self, template: BlankQuestionnaireTemplate) -> str:
        """Generate YAML with structured schema for blank template.
        
        Args:
            template: BlankQuestionnaireTemplate object with all data
            
        Returns:
            YAML string with structured template data
            
        Raises:
            TemplateGenerationError: If YAML generation fails
        """
        try:
            import yaml
            
            # Convert template to dictionary (same structure as JSON)
            template_dict = {
                "metadata": {
                    "template_version": template.metadata.template_version,
                    "baseline_version": template.metadata.baseline_version,
                    "export_date": template.metadata.export_date.isoformat(),
                    "total_control_count": template.metadata.total_control_count,
                    "frameworks_included": template.metadata.frameworks_included,
                    "instructions": template.metadata.instructions,
                },
                "controls": [
                    {
                        "id": control.id,
                        "title": control.title,
                        "description": control.description,
                        "family": control.family,
                        "in_moderate_baseline": control.in_moderate_baseline,
                    }
                    for control in template.controls
                ],
                "questions": {
                    control_id: [
                        {
                            "id": q.id,
                            "control_id": q.control_id,
                            "question_text": q.question_text,
                            "question_type": q.question_type.value,
                            "family": q.family,
                            "response": "",  # Blank response field
                        }
                        for q in questions
                    ]
                    for control_id, questions in template.questions.items()
                },
                "framework_mappings": {
                    control_id: {
                        "nist_csf": [
                            {
                                "function": m.function,
                                "category": m.category,
                                "subcategory": m.subcategory,
                            }
                            for m in mappings.nist_csf
                        ],
                        "glba": [
                            {
                                "requirement_type": m.requirement_type,
                                "requirement_id": m.requirement_id,
                                "description": m.description,
                            }
                            for m in mappings.glba
                        ],
                        "sox": [
                            {
                                "section": m.section,
                                "control_type": m.control_type,
                                "description": m.description,
                            }
                            for m in mappings.sox
                        ],
                        "ffiec": [
                            {
                                "domain": m.domain,
                                "assessment_factor": m.assessment_factor,
                                "description": m.description,
                            }
                            for m in mappings.ffiec
                        ],
                    }
                    for control_id, mappings in template.framework_mappings.items()
                },
                "aws_hints": template.aws_hints,
                "evidence": {
                    control.id: {
                        "description": "",
                        "location": "",
                        "notes": "",
                    }
                    for control in template.controls
                },
            }
            
            return yaml.dump(template_dict, default_flow_style=False, sort_keys=False)
            
        except Exception as e:
            raise TemplateGenerationError(f"Failed to generate YAML template: {str(e)}")
    
    # Template data preparation methods (Phase 2)
    def _collect_all_controls(self) -> List[Control]:
        """Collect all Moderate Baseline controls.
        
        Returns:
            List of all loaded controls
            
        Raises:
            TemplateNotReadyError: If controls not loaded
        """
        if self.controls is None:
            raise TemplateNotReadyError("Controls not loaded")
        return self.controls
    
    def _collect_all_questions(self) -> Dict[str, List[DiscoveryQuestion]]:
        """Collect all generated questions organized by control_id.
        
        Returns:
            Dictionary mapping control_id to list of questions
            
        Raises:
            TemplateNotReadyError: If questions not generated
        """
        if self.questions is None:
            raise TemplateNotReadyError("Questions not generated")
        return self.questions
    
    def _collect_all_mappings(self) -> Dict[str, FrameworkMappings]:
        """Collect all framework mappings organized by control_id.
        
        Returns:
            Dictionary mapping control_id to framework mappings
            
        Raises:
            TemplateNotReadyError: If mappings not available
        """
        if self.framework_mappings is None:
            raise TemplateNotReadyError("Framework mappings not available")
        return self.framework_mappings
    
    def _format_aws_hints(self, control_id: str) -> List[str]:
        """Format AWS hints as simple readable strings.
        
        Extracts AWS Config rules, Security Hub controls, and Control Tower IDs
        from framework mappings and formats them as simple strings.
        
        Args:
            control_id: NIST 800-53 control ID
            
        Returns:
            List of formatted AWS hint strings (empty list if no AWS mappings)
        """
        if self.framework_mappings is None or control_id not in self.framework_mappings:
            return []
        
        mappings = self.framework_mappings[control_id]
        hints = []
        
        for aws_control in mappings.aws:
            hint_parts = []
            
            # Add Config rules
            if aws_control.managed_controls.config_rules:
                config_rules = ", ".join(aws_control.managed_controls.config_rules)
                hint_parts.append(f"Config: {config_rules}")
            
            # Add Security Hub controls
            if aws_control.managed_controls.security_hub_controls:
                sh_controls = ", ".join(aws_control.managed_controls.security_hub_controls)
                hint_parts.append(f"Security Hub: {sh_controls}")
            
            # Add Control Tower IDs
            if aws_control.managed_controls.control_tower_ids:
                ct_ids = ", ".join(aws_control.managed_controls.control_tower_ids)
                hint_parts.append(f"Control Tower: {ct_ids}")
            
            if hint_parts:
                hints.append(", ".join(hint_parts))
        
        return hints
    
    def _build_template_metadata(self) -> TemplateMetadata:
        """Build template metadata with all required fields.
        
        Returns:
            TemplateMetadata object
            
        Raises:
            TemplateNotReadyError: If required data not available
        """
        self.is_ready_for_template_export()
        
        frameworks = ["NIST CSF", "GLBA", "SOX", "FFIEC"]
        if self.mcp_available:
            frameworks.append("AWS")
        
        instructions = self._generate_instructions()
        
        return TemplateMetadata(
            template_version=self.get_template_version(),
            baseline_version="NIST 800-53 Rev 5 Moderate Baseline",
            export_date=datetime.now(),
            total_control_count=len(self.controls) if self.controls else 0,
            frameworks_included=frameworks,
            instructions=instructions,
        )
    
    def _generate_instructions(self) -> str:
        """Generate instructions for completing the template.
        
        Returns:
            Formatted instructions text
        """
        instructions = """
# How to Complete This Questionnaire

## Overview
This questionnaire is based on the NIST 800-53 Revision 5 Moderate Baseline, which contains security controls appropriate for banking and financial institutions. Complete this assessment to evaluate your organization's compliance posture across multiple regulatory frameworks.

## Completing Response Fields

For each discovery question:
1. **Read the question carefully** - Questions are designed to elicit detailed narrative responses
2. **Provide specific details** - Avoid yes/no answers; describe your current implementation
3. **Be thorough** - Include information about processes, technologies, and responsible parties
4. **Note gaps** - If a control is not fully implemented, describe what's missing

## Documenting Evidence

For each control, document available evidence:
- **Evidence Description**: What the evidence is (e.g., "Access control policy document", "Audit log reports")
- **Evidence Location**: Where it's stored (e.g., "SharePoint site", "Compliance folder", "SIEM system")
- **Notes**: Any additional context or observations about the evidence

## Understanding Framework Mappings

This questionnaire maps NIST 800-53 controls to multiple frameworks:
- **NIST CSF**: Cybersecurity Framework functions and categories
- **GLBA**: Gramm-Leach-Bliley Act requirements
- **SOX**: Sarbanes-Oxley Act sections 302 and 404
- **FFIEC**: Federal Financial Institutions Examination Council standards

By completing the NIST 800-53 assessment, you automatically address requirements across all mapped frameworks.

## AWS Control Hints

Where applicable, AWS control hints are provided showing:
- **AWS Config Rules**: Automated compliance checks
- **Security Hub Controls**: Security findings and compliance status
- **Control Tower Controls**: Governance controls for AWS environments

These hints identify AWS services that can help implement or automate compliance controls.

## Question Types

Questions are organized by type to guide your assessment:
- **Current State**: Describe your existing implementation
- **Implementation**: Assess whether and how controls are implemented
- **Maturity**: Evaluate sophistication and effectiveness
- **Evidence**: Identify available documentation
- **Gap Analysis**: Identify deficiencies
- **Remediation**: Explore steps to achieve compliance
- **AWS Implementation**: Consider AWS services for compliance
- **Second Line Defense**: Assess readiness for compliance review
- **Third Line Defense**: Assess readiness for internal audit
- **Audit Readiness**: Evaluate automated reporting capabilities
- **Continuous Monitoring**: Assess automated monitoring and alerting

## Tips for Success

1. **Involve the right people** - Include IT, security, compliance, and business stakeholders
2. **Gather evidence first** - Collect relevant documentation before starting
3. **Be honest** - Accurate assessment is essential for effective remediation planning
4. **Focus on Moderate Baseline** - This assessment covers the most relevant controls for banking institutions
5. **Use AWS hints** - Consider cloud-based solutions for automation and efficiency

## Need Help?

Contact your compliance analyst or AWS solutions architect for assistance with this assessment.
"""
        return instructions.strip()
    
    def prepare_blank_template(self) -> BlankQuestionnaireTemplate:
        """Prepare a complete blank questionnaire template.
        
        Collects all controls, questions, mappings, and AWS hints,
        then builds a complete template ready for export.
        
        Returns:
            BlankQuestionnaireTemplate with all data
            
        Raises:
            TemplateNotReadyError: If required data not available
        """
        self.is_ready_for_template_export()
        
        controls = self._collect_all_controls()
        questions = self._collect_all_questions()
        mappings = self._collect_all_mappings()
        
        # Build AWS hints for all controls
        aws_hints = {}
        for control in controls:
            hints = self._format_aws_hints(control.id)
            if hints or not self.mcp_available:
                # Include entry even if empty when MCP unavailable (will add note in export)
                aws_hints[control.id] = hints
        
        metadata = self._build_template_metadata()
        
        return BlankQuestionnaireTemplate(
            controls=controls,
            questions=questions,
            framework_mappings=mappings,
            aws_hints=aws_hints,
            metadata=metadata,
        )
