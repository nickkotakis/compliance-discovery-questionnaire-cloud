"""Helper functions for Excel template generation."""

from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill
from openpyxl.worksheet.worksheet import Worksheet


def create_header_style():
    """Create style for header rows.
    
    Returns:
        Tuple of (font, fill, alignment)
    """
    font = Font(bold=True, size=11)
    fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    return font, fill, alignment


def apply_header_formatting(ws: Worksheet, headers: list, row: int = 1):
    """Apply formatting to header row.
    
    Args:
        ws: Worksheet to format
        headers: List of header texts
        row: Row number for headers (default 1)
    """
    font, fill, alignment = create_header_style()
    
    for col_idx, header in enumerate(headers, start=1):
        cell = ws.cell(row=row, column=col_idx, value=header)
        cell.font = Font(bold=True, size=11, color="FFFFFF")
        cell.fill = fill
        cell.alignment = alignment


def set_column_widths(ws: Worksheet, widths: dict):
    """Set column widths.
    
    Args:
        ws: Worksheet to format
        widths: Dictionary mapping column letter to width
    """
    for col, width in widths.items():
        ws.column_dimensions[col].width = width


def freeze_header_row(ws: Worksheet):
    """Freeze the first row (header row).
    
    Args:
        ws: Worksheet to freeze
    """
    ws.freeze_panes = "A2"


def enable_text_wrap(ws: Worksheet, columns: list):
    """Enable text wrapping for specified columns.
    
    Args:
        ws: Worksheet to format
        columns: List of column letters to wrap
    """
    for col in columns:
        for cell in ws[col]:
            if cell.row > 1:  # Skip header
                cell.alignment = Alignment(wrap_text=True, vertical="top")
